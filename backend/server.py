from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import json
import logging
import uuid
import secrets
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from typing import Optional
from fastapi import FastAPI, APIRouter, Request, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
import math
import re
import requests as http_requests
import asyncio
import smtplib
import zipfile
import base64
import io
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# Sold-out items reset daily at midnight IST; helpers below.
IST = ZoneInfo("Asia/Kolkata")

def today_ist_str() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d")

# ── Config ──────────────────────────────────────────────────────────────
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ.get("DB_NAME", "perfectly_good")
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE = timedelta(hours=24)
REFRESH_TOKEN_EXPIRE = timedelta(days=7)
RAZORPAY_KEY_ID = os.environ.get("RAZORPAY_KEY_ID", "rzp_test_SSfFeyx6ytVg0B")
RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "")
GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ── Razorpay client ─────────────────────────────────────────────────────
import razorpay
razorpay_client = None
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    try:
        razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
        logger.info(f"Razorpay client initialized ({'LIVE' if RAZORPAY_KEY_ID.startswith('rzp_live') else 'TEST'} mode)")
    except Exception as e:
        logger.error(f"Razorpay client init failed: {e}")


# ── MongoDB ─────────────────────────────────────────────────────────────
mongo_client = AsyncIOMotorClient(MONGO_URL)
db = mongo_client[DB_NAME]

# ── Helpers ─────────────────────────────────────────────────────────────
def gen_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + ACCESS_TOKEN_EXPIRE,
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + REFRESH_TOKEN_EXPIRE,
        "type": "refresh",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def user_response(user: dict) -> dict:
    role = user.get("role", "user")
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user.get("name", ""),
        "phone": user.get("phone", ""),
        "role": role,
        "permissions": sorted(get_effective_permissions(role, user.get("permission_overrides"))),
        "permission_overrides": user.get("permission_overrides") or {},
        "picture": user.get("picture"),
        "location": user.get("location"),
        "parent_vendor_id": user.get("parent_vendor_id"),
        "staff_permissions": user.get("staff_permissions") or [],
        "created_at": user.get("created_at", datetime.now(timezone.utc)).isoformat(),
    }

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


VENDOR_ROLES = ("vendor", "admin", "vendor_staff")

async def _resolve_vendor(user: dict):
    """Return the vendor doc for a vendor/admin owner OR a vendor_staff member."""
    if user.get("role") == "vendor_staff":
        vid = user.get("parent_vendor_id")
        return await db.vendors.find_one({"vendor_id": vid}, {"_id": 0}) if vid else None
    return await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})

def _require_staff_perm(user: dict, perm: str):
    """Owners (vendor/admin) may do anything; vendor_staff need the specific permission."""
    role = user.get("role")
    if role in ("vendor", "admin"):
        return
    if role == "vendor_staff" and perm in (user.get("staff_permissions") or []):
        return
    raise HTTPException(status_code=403, detail="You do not have permission for this action")

# ── RBAC ────────────────────────────────────────────────────────────────
STAFF_ROLES = {"admin", "operations", "customer_success", "finance"}

PERMISSIONS = [
    "view_dashboard", "view_vendors", "manage_vendors",
    "manage_menu", "upload_images", "ai_import",
    "view_orders", "update_order_status", "manage_orders",
    "view_finance", "manage_payouts",
    "view_users", "manage_users",
    "manage_settings", "manage_roles", "add_notes",
    "manage_support",
]

ROLE_PERMISSIONS = {
    "user": set(),
    "vendor": {"manage_menu", "view_orders", "update_order_status", "view_finance"},
    "admin": set(PERMISSIONS),
    "operations": {
        "view_dashboard", "view_vendors", "manage_vendors", "manage_menu",
        "upload_images", "ai_import", "view_orders", "update_order_status",
        "manage_orders", "add_notes", "manage_support",
    },
    "customer_success": {
        "view_dashboard", "view_vendors", "view_orders", "update_order_status", "add_notes",
    },
    "finance": {
        "view_dashboard", "view_vendors", "view_finance", "view_orders",
        "manage_payouts",
    },
}


def get_effective_permissions(role: str, overrides: Optional[dict] = None) -> set:
    perms = set(ROLE_PERMISSIONS.get(role, set()))
    if overrides:
        for perm, allowed in overrides.items():
            if allowed:
                perms.add(perm)
            else:
                perms.discard(perm)
    return perms


async def require_permission(request: Request, permission: str) -> dict:
    """Authenticate and ensure the current user is a staff member with the given permission.
    Ops routes are staff-only: vendors/customers can never access them even if a role
    happens to share a permission name."""
    user = await get_current_user(request)
    if user.get("role") not in STAFF_ROLES:
        raise HTTPException(status_code=403, detail="Staff access only")
    perms = get_effective_permissions(user.get("role", "user"), user.get("permission_overrides"))
    if permission not in perms:
        raise HTTPException(status_code=403, detail=f"Insufficient permissions (requires '{permission}')")
    return user


# ── Platform settings ───────────────────────────────────────────────────
DEFAULT_SETTINGS = {
    "commission_rate": 0.15,
    "gst_on_commission": 0.18,
    "gst_rate": 0.05,
    "convenience_rate": 0.05,
    "default_discount_pct": 40,
    "categories": ["Bakery", "Restaurant", "Cafe", "Grocery", "QSR", "Cloud Kitchen", "Dessert"],
    "pickup_slots": ["12:00-15:00", "15:00-18:00", "17:00-20:00", "18:00-21:00", "19:00-22:00"],
    "service_types": ["takeaway", "dine_in", "both"],
}


async def get_settings_doc() -> dict:
    doc = await db.settings.find_one({"_id": "platform"})
    if not doc:
        doc = {"_id": "platform", **DEFAULT_SETTINGS}
        await db.settings.insert_one(doc)
        return doc
    changed = {}
    for k, v in DEFAULT_SETTINGS.items():
        if k not in doc:
            changed[k] = v
    if changed:
        await db.settings.update_one({"_id": "platform"}, {"$set": changed})
        doc.update(changed)
    return doc


def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))

def resolve_place_id(place_id: str) -> dict:
    """Kept for backwards compat — not used when no Google API key."""
    return {}

def geocode_address(address: str) -> dict:
    """Geocode an address using free OpenStreetMap Nominatim API."""
    if not address:
        return {}
    try:
        resp = http_requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address, "format": "json", "limit": 1, "addressdetails": 1},
            headers={"User-Agent": "PerfectlyGood/1.0"},
            timeout=5,
        )
        results = resp.json()
        if results and len(results) > 0:
            r = results[0]
            lat = float(r.get("lat", 0))
            lon = float(r.get("lon", 0))
            display = r.get("display_name", address)
            return {
                "lat": lat,
                "lon": lon,
                "address": display,
                "maps_url": f"https://www.google.com/maps/search/?api=1&query={lat},{lon}",
            }
        logger.warning(f"Geocode failed for: {address}")
    except Exception as e:
        logger.error(f"Geocode error: {e}")
    return {}

# ── App ─────────────────────────────────────────────────────────────────
app = FastAPI()
api = APIRouter(prefix="/api")

# ══════════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════════

class RegisterBody(BaseModel):
    name: str
    email: str
    phone: str
    password: str

class LoginBody(BaseModel):
    email: str
    password: str

class ForgotPasswordBody(BaseModel):
    email: str

class ResetPasswordBody(BaseModel):
    token: str
    new_password: str

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

@api.post("/auth/register")
async def register(body: RegisterBody):
    email = body.email.strip().lower()
    if not EMAIL_REGEX.match(email):
        raise HTTPException(status_code=400, detail="Please enter a valid email address")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="Name is required")
    phone = (body.phone or "").strip()
    if len(re.sub(r"\D", "", phone)) < 10:
        raise HTTPException(status_code=400, detail="Please enter a valid phone number")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = gen_id("user")
    user_doc = {
        "user_id": user_id,
        "email": email,
        "name": body.name.strip(),
        "phone": phone,
        "password_hash": hash_password(body.password),
        "role": "user",
        "picture": None,
        "location": None,
        "created_at": datetime.now(timezone.utc),
    }
    await db.users.insert_one(user_doc)
    token = create_access_token(user_id, email)
    resp = user_response(user_doc)
    resp["access_token"] = token
    response = JSONResponse(content=resp)
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=86400, path="/")
    return response

@api.post("/auth/login")
async def login(body: LoginBody):
    email = body.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["user_id"], email)
    resp = user_response(user)
    resp["access_token"] = token
    response = JSONResponse(content=resp)
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=86400, path="/")
    return response

# ── Sign in with Apple (native iOS) ─────────────────────────────────────
APPLE_ISSUER = "https://appleid.apple.com"
APPLE_JWKS_URL = "https://appleid.apple.com/auth/keys"
# Must include the app bundle id AND host.exp.Exponent (Expo Go).
APPLE_AUDIENCES = [a.strip() for a in os.environ.get(
    "APPLE_AUDIENCES", "in.perfectlygood.app,host.exp.Exponent"
).split(",") if a.strip()]
_apple_jwks = jwt.PyJWKClient(APPLE_JWKS_URL)


class AppleAuthBody(BaseModel):
    identity_token: str
    name: Optional[str] = None
    email: Optional[str] = None


@api.post("/auth/apple")
async def apple_auth(body: AppleAuthBody):
    """Verify an Apple identity token against Apple's JWKS and issue our JWT.
    Users are keyed by the Apple `sub` (name/email arrive only on first sign-in)."""
    try:
        signing_key = _apple_jwks.get_signing_key_from_jwt(body.identity_token)
        claims = jwt.decode(
            body.identity_token,
            signing_key.key,
            algorithms=["RS256"],
            audience=APPLE_AUDIENCES,
            issuer=APPLE_ISSUER,
        )
    except Exception as e:
        logger.warning(f"Apple token verification failed: {e}")
        raise HTTPException(status_code=401, detail="Invalid Apple identity token")

    apple_sub = claims.get("sub")
    if not apple_sub:
        raise HTTPException(status_code=401, detail="Invalid Apple identity token")

    email = (body.email or claims.get("email") or "").strip().lower()
    now = datetime.now(timezone.utc)

    user = await db.users.find_one({"apple_sub": apple_sub}, {"_id": 0})
    if not user and email:
        user = await db.users.find_one({"email": email}, {"_id": 0})

    if not user:
        user_id = gen_id("user")
        name = (body.name or "").strip() or (email.split("@")[0] if email else "Apple User")
        user = {
            "user_id": user_id,
            "email": email or f"{apple_sub}@privaterelay.appleid.com",
            "name": name,
            "phone": "",
            "role": "user",
            "apple_sub": apple_sub,
            "password_hash": None,
            "permission_overrides": {},
            "picture": None,
            "location": None,
            "created_at": now,
        }
        await db.users.insert_one(user)
    elif not user.get("apple_sub"):
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"apple_sub": apple_sub}})

    token = create_access_token(user["user_id"], user["email"])
    resp = user_response(user)
    resp["access_token"] = token
    response = JSONResponse(content=resp)
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=86400, path="/")
    return response

@api.get("/auth/me")
async def auth_me(request: Request):
    user = await get_current_user(request)
    return user_response(user)

@api.delete("/auth/me")
async def delete_my_account(request: Request):
    """In-app account deletion (Apple 5.1.1(v)). Removes the user's account and
    associated personal data. Orders are anonymised so vendor/payout records stay intact."""
    user = await get_current_user(request)
    if user.get("role") in STAFF_ROLES:
        raise HTTPException(status_code=403, detail="Staff accounts cannot be deleted from the app.")
    uid = user["user_id"]
    await db.support_requests.delete_many({"user_id": uid})
    await db.pending_orders.delete_many({"user_id": uid})
    await db.password_reset_tokens.delete_many({"user_id": uid})
    await db.deal_alerts.delete_many({"user_id": uid})
    await db.orders.update_many({"user_id": uid}, {"$set": {"user_name": "Deleted user"}})
    # If this user owns a vendor profile, detach it (kept for records).
    await db.vendors.update_many({"user_id": uid}, {"$set": {"user_id": None}})
    await db.users.delete_one({"user_id": uid})
    response = JSONResponse(content={"message": "Account deleted"})
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

@api.post("/auth/logout")
async def auth_logout():
    response = JSONResponse(content={"message": "Logged out"})
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordBody):
    email = body.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Email not found")
    reset_token = secrets.token_urlsafe(32)
    await db.password_reset_tokens.insert_one({
        "token": reset_token,
        "user_id": user["user_id"],
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
        "used": False,
    })
    logger.info(f"Password reset token for {email}: {reset_token}")
    return {"reset_token": reset_token}

@api.post("/auth/reset-password")
async def reset_password(body: ResetPasswordBody):
    record = await db.password_reset_tokens.find_one({"token": body.token, "used": False})
    if not record:
        raise HTTPException(status_code=400, detail="Invalid or expired token")
    if record["expires_at"].replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Token expired")
    await db.users.update_one(
        {"user_id": record["user_id"]},
        {"$set": {"password_hash": hash_password(body.new_password)}},
    )
    await db.password_reset_tokens.update_one({"token": body.token}, {"$set": {"used": True}})
    return {"message": "Password reset successfully"}

@api.post("/auth/change-password")
async def change_password(body: dict, request: Request):
    """Logged-in user changes their own password (requires current password)."""
    user = await get_current_user(request)
    current = (body.get("current_password") or "").strip()
    new_password = (body.get("new_password") or "").strip()
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    full = await db.users.find_one({"user_id": user["user_id"]})
    if not full or not verify_password(current, full.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"password_hash": hash_password(new_password)}})
    return {"message": "Password changed successfully"}

# ── Push Notifications ──────────────────────────────────────────────────

class PushTokenBody(BaseModel):
    push_token: str

@api.post("/auth/push-token")
async def save_push_token(body: PushTokenBody, request: Request):
    user = await get_current_user(request)
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"push_token": body.push_token}},
    )
    return {"message": "Push token saved"}

async def send_push_to_vendor(vendor_id: str, title: str, body: str, data: dict = None):
    """Send push notification to all devices registered for this vendor."""
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        return
    vendor_user = await db.users.find_one({"user_id": vendor.get("user_id")}, {"_id": 0})
    if not vendor_user or not vendor_user.get("push_token"):
        logger.info(f"No push token for vendor {vendor_id}, skipping notification")
        return

    push_token = vendor_user["push_token"]
    message = {
        "to": push_token,
        "sound": "default",
        "title": title,
        "body": body,
        "channelId": "orders",
    }
    if data:
        message["data"] = data

    try:
        resp = http_requests.post(
            "https://exp.host/--/api/v2/push/send",
            json=message,
            headers={"Content-Type": "application/json"},
            timeout=5,
        )
        logger.info(f"Push sent to vendor {vendor_id}: {resp.status_code}")
    except Exception as e:
        logger.error(f"Push notification failed: {e}")

async def send_push_to_all_users(title: str, body: str, data: dict = None):
    """Broadcast push notification to all users (non-vendor) with push tokens."""
    users = await db.users.find(
        {"push_token": {"$exists": True, "$ne": None}, "role": "user"},
        {"_id": 0, "push_token": 1},
    ).to_list(10000)
    tokens = [u["push_token"] for u in users if u.get("push_token")]
    if not tokens:
        logger.info("No user push tokens found, skipping broadcast")
        return

    messages = []
    for token in tokens:
        msg = {"to": token, "sound": "default", "title": title, "body": body, "channelId": "orders"}
        if data:
            msg["data"] = data
        messages.append(msg)

    # Expo push API supports batches of up to 100
    for i in range(0, len(messages), 100):
        batch = messages[i:i+100]
        try:
            resp = http_requests.post(
                "https://exp.host/--/api/v2/push/send",
                json=batch,
                headers={"Content-Type": "application/json"},
                timeout=10,
            )
            logger.info(f"Broadcast push sent to {len(batch)} users: {resp.status_code}")
        except Exception as e:
            logger.error(f"Broadcast push failed: {e}")

# ══════════════════════════════════════════════════════════════════════════
#  DROPS
# ══════════════════════════════════════════════════════════════════════════

@api.get("/drops/categories")
async def get_categories():
    categories = await db.vendors.distinct("category")
    return categories or []

def item_to_drop(item: dict, vendor: Optional[dict]) -> dict:
    """Shape a menu_item (+ its vendor) into the legacy drop response used by the app."""
    out = dict(item)
    out["item_id"] = item.get("menu_item_id")
    out["is_active"] = bool(item.get("available_today"))
    if vendor:
        out["vendor_name"] = vendor.get("name", "")
        out["vendor_location"] = vendor.get("location", {})
        out["vendor_category"] = vendor.get("category", "")
        out["pickup_start_time"] = vendor.get("pickup_start_time", "")
        out["pickup_end_time"] = vendor.get("pickup_end_time", "")
        out["service_type"] = vendor.get("service_type", "both")
        _ver = vendor.get("verification") or {}
        _agr = _ver.get("agreement") or {}
        out["vendor_verified"] = vendor.get("status") == "active" and bool(_agr.get("accepted")) and bool((_ver.get("fssai_number") or "").strip())
    for k in ("created_at", "updated_at"):
        if k in out and hasattr(out[k], "isoformat"):
            out[k] = out[k].isoformat()
    return out


@api.get("/drops")
async def list_drops(
    lat: Optional[float] = None,
    lon: Optional[float] = None,
    search: Optional[str] = None,
    category: Optional[str] = None,
    max_price: Optional[float] = None,
    sort_by: Optional[str] = None,
):
    # Only items toggled available today, from active vendors
    active_vendor_ids = await db.vendors.distinct("vendor_id", {"status": "active"})
    query: dict = {"available_today": True, "in_stock": {"$ne": False}, "vendor_id": {"$in": active_vendor_ids}}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
        ]
    if category:
        cat_vendor_ids = await db.vendors.distinct("vendor_id", {"category": category})
        query["vendor_id"] = {"$in": [v for v in active_vendor_ids if v in set(cat_vendor_ids)]}
    if max_price:
        query["discounted_price"] = {"$lte": max_price}

    items = await db.menu_items.find(query, {"_id": 0}).to_list(500)
    vendor_ids = list(set(i.get("vendor_id") for i in items if i.get("vendor_id")))
    vendors_list = await db.vendors.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0}).to_list(500) if vendor_ids else []
    vendor_cache = {v["vendor_id"]: v for v in vendors_list}
    drops = [item_to_drop(i, vendor_cache.get(i.get("vendor_id"))) for i in items]

    if sort_by == "price":
        drops.sort(key=lambda d: d.get("discounted_price", 0))
    elif sort_by == "discount":
        drops.sort(key=lambda d: (d.get("original_price", 1) - d.get("discounted_price", 0)) / max(d.get("original_price", 1), 1), reverse=True)
    elif lat and lon:
        for d in drops:
            vloc = d.get("vendor_location", {})
            d["_dist"] = haversine(lat, lon, vloc["lat"], vloc["lon"]) if vloc and vloc.get("lat") and vloc.get("lon") else 99999
        drops.sort(key=lambda d: d.get("_dist", 99999))
        for d in drops:
            d.pop("_dist", None)

    return drops

@api.get("/drops/{item_id}")
async def get_drop(item_id: str, lat: Optional[float] = None, lon: Optional[float] = None):
    item = await db.menu_items.find_one({"menu_item_id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    vendor = await db.vendors.find_one({"vendor_id": item.get("vendor_id")}, {"_id": 0})
    return item_to_drop(item, vendor)

# ══════════════════════════════════════════════════════════════════════════
#  RESTAURANTS  (customer browse: surplus + takeaway + dine-in)
# ══════════════════════════════════════════════════════════════════════════

def _vendor_public(v: dict) -> dict:
    ver = v.get("verification") or {}
    agreement = ver.get("agreement") or {}
    return {
        "vendor_id": v.get("vendor_id"),
        "name": v.get("name", ""),
        "category": v.get("category", ""),
        "service_type": v.get("service_type", "both"),
        "location": v.get("location", {}),
        "logo_url": v.get("logo_url", ""),
        "storefront_image": v.get("storefront_image", ""),
        "discount_percentage": v.get("discount_percentage", 0) or 0,
        "pickup_start_time": v.get("pickup_start_time", ""),
        "pickup_end_time": v.get("pickup_end_time", ""),
        # Verified = admin-approved (active) AND has completed compliance (FSSAI + accepted agreement).
        "verified": v.get("status") == "active" and bool(agreement.get("accepted")) and bool((ver.get("fssai_number") or "").strip()),
    }

async def _ops_name_map() -> dict:
    """Map operations staff user_id -> display name (for vendor assignment)."""
    ops = await db.users.find({"role": "operations"}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).to_list(500)
    return {o["user_id"]: (o.get("name") or o.get("email") or "") for o in ops}

def _menu_public(m: dict, order_type: str, discount_pct: float = 0) -> dict:
    """Shape a menu_item for the customer. Surplus uses discounted_price; takeaway/dine-in use the menu (original) price with the vendor's flat discount applied."""
    op = m.get("original_price") or 0
    dp = m.get("discounted_price") or 0
    is_surplus = order_type == "surplus"
    disc = discount_pct or 0
    takeaway_price = round(op * (1 - disc / 100), 2) if op else 0
    out = {
        "menu_item_id": m.get("menu_item_id"),
        "item_id": m.get("menu_item_id"),
        "name": m.get("name", ""),
        "description": m.get("description", ""),
        "image_url": m.get("image_url", ""),
        "food_type": m.get("food_type", "veg"),
        "contains_egg": bool(m.get("contains_egg")),
        "serving_size": m.get("serving_size", ""),
        "category": m.get("category", ""),
        "kcal": m.get("kcal"),
        "protein": m.get("protein"),
        "original_price": op,
        "price": dp if is_surplus else takeaway_price,
        "available_today": bool(m.get("available_today")),
    }
    if is_surplus:
        out["discounted_price"] = dp
        out["quantity_available"] = m.get("quantity_available")
        out["expiry"] = m.get("expiry", "")
        out["discount"] = round(((op - dp) / op) * 100) if op else 0
    else:
        out["discount_percentage"] = disc
    return out

@api.get("/restaurants")
async def list_restaurants(
    lat: Optional[float] = None,
    lon: Optional[float] = None,
    search: Optional[str] = None,
    category: Optional[str] = None,
):
    query: dict = {"status": "active"}
    if category:
        query["category"] = category
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    vendors = await db.vendors.find(query, {"_id": 0}).to_list(500)
    vendor_ids = [v["vendor_id"] for v in vendors]

    counts: dict = {}
    if vendor_ids:
        pipeline = [
            {"$match": {"vendor_id": {"$in": vendor_ids}}},
            {"$group": {
                "_id": "$vendor_id",
                "menu_count": {"$sum": 1},
                "surplus_count": {"$sum": {"$cond": [{"$eq": ["$available_today", True]}, 1, 0]}},
            }},
        ]
        async for row in db.menu_items.aggregate(pipeline):
            counts[row["_id"]] = {"menu_count": row["menu_count"], "surplus_count": row["surplus_count"]}

    out = []
    for v in vendors:
        pub = _vendor_public(v)
        c = counts.get(v["vendor_id"], {})
        pub["menu_count"] = c.get("menu_count", 0)
        pub["surplus_count"] = c.get("surplus_count", 0)
        loc = v.get("location", {}) or {}
        if lat is not None and lon is not None and loc.get("lat") and loc.get("lon"):
            pub["distance"] = round(haversine(lat, lon, loc["lat"], loc["lon"]), 1)
        else:
            pub["distance"] = None
        out.append(pub)

    # Restaurants with active surplus deals first, then nearest
    out.sort(key=lambda r: (0 if r["surplus_count"] > 0 else 1, r["distance"] if r["distance"] is not None else 99999))
    return out

@api.get("/restaurants/{vendor_id}")
async def get_restaurant(vendor_id: str):
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v or v.get("status") != "active":
        raise HTTPException(status_code=404, detail="Restaurant not found")
    items = await db.menu_items.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(1000)
    in_stock = lambda m: m.get("in_stock", True) is not False
    disc = v.get("discount_percentage") or 0
    surplus_items = [_menu_public(m, "surplus", disc) for m in items if m.get("available_today") and in_stock(m)]
    menu_items = [_menu_public(m, "takeaway", disc) for m in items if in_stock(m)]
    return {
        "vendor": _vendor_public(v),
        "surplus_items": surplus_items,
        "menu_items": menu_items,
    }

# ══════════════════════════════════════════════════════════════════════════
#  ORDERS
# ══════════════════════════════════════════════════════════════════════════

class CreateOrderBody(BaseModel):
    food_item_id: str
    quantity: int
    order_type: Optional[str] = "surplus"  # "surplus" | "takeaway" | "dine_in"

class VerifyOrderBody(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    food_item_id: str
    quantity: int
    order_type: Optional[str] = "surplus"

@api.post("/orders/create")
async def create_order(body: CreateOrderBody, request: Request):
    user = await get_current_user(request)
    order_type = body.order_type if body.order_type in ("surplus", "takeaway", "dine_in") else "surplus"

    if order_type == "surplus":
        item = await db.menu_items.find_one({"menu_item_id": body.food_item_id, "available_today": True}, {"_id": 0})
        if not item:
            raise HTTPException(status_code=404, detail="Item not available")
        if item.get("in_stock") is False:
            raise HTTPException(status_code=400, detail="This item is sold out")
        qty_avail = item.get("quantity_available")
        if qty_avail is not None and body.quantity > qty_avail:
            raise HTTPException(status_code=400, detail="Not enough quantity available")
        unit_price = item.get("discounted_price") or 0
    else:
        # Takeaway / Dine-in: order from the regular menu at menu (original) price
        # with the vendor's flat discount applied.
        item = await db.menu_items.find_one({"menu_item_id": body.food_item_id}, {"_id": 0})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        if item.get("in_stock") is False:
            raise HTTPException(status_code=400, detail="This item is sold out")
        vendor = await db.vendors.find_one({"vendor_id": item.get("vendor_id")}, {"_id": 0})
        disc = (vendor or {}).get("discount_percentage") or 0
        base_price = item.get("original_price") or item.get("discounted_price") or 0
        unit_price = round(base_price * (1 - disc / 100), 2)

    # Vendor must be Active (admin-approved) to receive orders.
    _ov = await db.vendors.find_one({"vendor_id": item.get("vendor_id")}, {"_id": 0})
    if not _ov or _ov.get("status") != "active":
        raise HTTPException(status_code=400, detail="This restaurant is not currently accepting orders.")

    cfg = await get_settings_doc()
    subtotal = round(unit_price * body.quantity, 2)
    gst = round(subtotal * cfg.get("gst_rate", 0.05), 2)
    convenience_fee = round(subtotal * cfg.get("convenience_rate", 0.05), 2)
    total = round(subtotal + gst + convenience_fee, 2)
    amount_paise = int(round(total, 2) * 100)

    # Create a real Razorpay order so the checkout can process the payment
    if not razorpay_client:
        raise HTTPException(status_code=503, detail="Payments are not configured. Please try again later.")
    try:
        rzp_order = razorpay_client.order.create({
            "amount": amount_paise,
            "currency": "INR",
            "receipt": f"rcpt_{secrets.token_hex(6)}",  # must be <= 40 chars
            "payment_capture": 1,
        })
    except Exception as e:
        logger.error(f"Razorpay order create failed: {e}")
        raise HTTPException(status_code=502, detail="Could not start payment. Please try again.")

    razorpay_order_id = rzp_order["id"]

    await db.pending_orders.insert_one({
        "razorpay_order_id": razorpay_order_id,
        "user_id": user["user_id"],
        "food_item_id": body.food_item_id,
        "quantity": body.quantity,
        "order_type": order_type,
        "unit_price": round(unit_price, 2),
        "item_subtotal": round(subtotal, 2),
        "total_amount": total,
        "created_at": datetime.now(timezone.utc),
    })

    return {
        "razorpay_order_id": razorpay_order_id,
        "key_id": RAZORPAY_KEY_ID,
        "amount": amount_paise,
    }

async def _gen_pickup_code() -> str:
    """Unique 6-digit numeric pickup code, never reused across orders."""
    for _ in range(25):
        code = f"{secrets.randbelow(900000) + 100000}"
        if not await db.orders.find_one({"pickup_code": code}, {"_id": 1}):
            return code
    # Extremely unlikely fallback
    return f"{secrets.randbelow(900000) + 100000}"


async def _finalize_order(pending: dict, razorpay_payment_id: str = "") -> Optional[str]:
    """Create the confirmed order from a pending order. Idempotent: if an order
    for this razorpay_order_id already exists, it is returned without duplication.
    Used by both /orders/verify (client callback) and the Razorpay webhook."""
    rzp_order_id = pending.get("razorpay_order_id")

    # Idempotency guard — never create two orders for the same payment
    existing = await db.orders.find_one({"razorpay_order_id": rzp_order_id}, {"_id": 0})
    if existing:
        await db.pending_orders.delete_one({"razorpay_order_id": rzp_order_id})
        return existing.get("order_id")

    item = await db.menu_items.find_one({"menu_item_id": pending.get("food_item_id")}, {"_id": 0})
    if not item:
        return None
    vendor = await db.vendors.find_one({"vendor_id": item.get("vendor_id")}, {"_id": 0})
    order_user = await db.users.find_one({"user_id": pending.get("user_id")}, {"_id": 0}) or {}

    order_id = gen_id("order")
    order_type = pending.get("order_type", "surplus")
    quantity = pending.get("quantity", 1)
    unit_price = pending.get("unit_price", item.get("discounted_price", 0))
    now = datetime.now(timezone.utc)
    pickup_code = await _gen_pickup_code()
    order_doc = {
        "order_id": order_id,
        "user_id": pending.get("user_id"),
        "user_name": order_user.get("name", ""),
        "food_item_id": pending.get("food_item_id"),
        "food_item_name": item.get("name", ""),
        "vendor_id": item.get("vendor_id", ""),
        "vendor_name": vendor.get("name", "") if vendor else "",
        "quantity": quantity,
        "order_type": order_type,
        "discounted_price": unit_price,
        "item_subtotal": pending.get("item_subtotal", round(unit_price * quantity, 2)),
        "total_amount": pending.get("total_amount", 0),
        "status": "reserved",
        "pickup_code": pickup_code,
        "pickup_verified": False,
        "pickup_verified_at": None,
        "pickup_verified_by": None,
        "payment_confirmed_at": now,
        "pickup_start_time": vendor.get("pickup_start_time", "") if vendor else "",
        "pickup_end_time": vendor.get("pickup_end_time", "") if vendor else "",
        "razorpay_order_id": rzp_order_id,
        "razorpay_payment_id": razorpay_payment_id,
        "created_at": now,
    }
    await db.orders.insert_one(order_doc)

    # Decrement available quantity for surplus listings only (if tracked)
    if order_type == "surplus" and item.get("quantity_available") is not None:
        await db.menu_items.update_one(
            {"menu_item_id": pending.get("food_item_id")},
            {"$inc": {"quantity_available": -quantity}},
        )
    await db.vendors.update_one({"vendor_id": item.get("vendor_id", "")}, {"$set": {"last_order_date": datetime.now(timezone.utc)}})
    await db.pending_orders.delete_one({"razorpay_order_id": rzp_order_id})

    # Send push notification to vendor
    await send_push_to_vendor(
        vendor_id=item.get("vendor_id", ""),
        title="New Order!",
        body=f"{order_user.get('name', 'A customer')} reserved {quantity}× {item.get('name', 'item')} — ₹{pending.get('total_amount', 0)}",
        data={"order_id": order_id, "type": "new_order"},
    )
    return order_id


@api.post("/orders/verify")
async def verify_order(body: VerifyOrderBody, request: Request):
    user = await get_current_user(request)

    async def _order_out(oid: str) -> dict:
        o = await db.orders.find_one({"order_id": oid}, {"_id": 0})
        if o and isinstance(o.get("created_at"), datetime):
            o["created_at"] = o["created_at"].isoformat()
        if o and isinstance(o.get("payment_confirmed_at"), datetime):
            o["payment_confirmed_at"] = o["payment_confirmed_at"].isoformat()
        return o or {}

    pending = await db.pending_orders.find_one({"razorpay_order_id": body.razorpay_order_id}, {"_id": 0})
    if not pending:
        # Webhook may have already finalized this order
        existing = await db.orders.find_one({"razorpay_order_id": body.razorpay_order_id}, {"_id": 0})
        if existing:
            return {"message": "Order confirmed", "order_id": existing.get("order_id"), "order": await _order_out(existing.get("order_id"))}
        raise HTTPException(status_code=400, detail="Order not found")

    # Verify the Razorpay payment signature (HMAC SHA256 of order_id|payment_id)
    if not razorpay_client:
        raise HTTPException(status_code=503, detail="Payments are not configured.")
    try:
        razorpay_client.utility.verify_payment_signature({
            "razorpay_order_id": body.razorpay_order_id,
            "razorpay_payment_id": body.razorpay_payment_id,
            "razorpay_signature": body.razorpay_signature,
        })
    except Exception as e:
        logger.warning(f"Razorpay signature verification failed: {e}")
        raise HTTPException(status_code=400, detail="Payment verification failed")

    order_id = await _finalize_order(pending, body.razorpay_payment_id)
    if not order_id:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Order confirmed", "order_id": order_id, "order": await _order_out(order_id)}


@api.post("/webhooks/razorpay")
async def razorpay_webhook(request: Request):
    """Server-to-server confirmation from Razorpay. Reliably finalizes an order
    even if the app is closed before /orders/verify runs. Configure this URL and
    a secret in the Razorpay Dashboard (events: payment.captured, order.paid, payment.failed, refund.processed)."""
    raw = await request.body()
    signature = request.headers.get("X-Razorpay-Signature", "")
    webhook_secret = os.environ.get("RAZORPAY_WEBHOOK_SECRET", "")

    if not webhook_secret or not razorpay_client:
        logger.error("Razorpay webhook received but RAZORPAY_WEBHOOK_SECRET is not set")
        raise HTTPException(status_code=503, detail="Webhook not configured")

    try:
        razorpay_client.utility.verify_webhook_signature(raw.decode("utf-8"), signature, webhook_secret)
    except Exception as e:
        logger.warning(f"Razorpay webhook signature verification failed: {e}")
        raise HTTPException(status_code=400, detail="Invalid signature")

    try:
        payload = json.loads(raw.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid payload")

    event = payload.get("event", "")

    # Handle failed payments: log the attempt and drop the stale pending order so
    # it doesn't linger. No confirmed order is ever created for a failed payment.
    if event == "payment.failed":
        try:
            entity = payload["payload"]["payment"]["entity"]
        except Exception:
            entity = {}
        rzp_order_id = entity.get("order_id", "")
        await db.payment_failures.insert_one({
            "razorpay_order_id": rzp_order_id,
            "razorpay_payment_id": entity.get("id", ""),
            "amount": (entity.get("amount") or 0) / 100,
            "error_code": entity.get("error_code", ""),
            "error_description": entity.get("error_description", ""),
            "method": entity.get("method", ""),
            "created_at": datetime.now(timezone.utc),
        })
        if rzp_order_id:
            await db.pending_orders.delete_one({"razorpay_order_id": rzp_order_id})
        logger.info(f"Webhook: payment.failed for {rzp_order_id} ({entity.get('error_description', 'no detail')})")
        return {"status": "failed_logged"}

    # Handle refunds: mark the matching order refunded and invalidate its pickup code.
    if event in ("refund.created", "refund.processed", "refund.speed_changed"):
        try:
            refund_entity = payload["payload"]["refund"]["entity"]
        except Exception:
            refund_entity = {}
        payment_id = refund_entity.get("payment_id", "")
        if payment_id:
            order = await db.orders.find_one({"razorpay_payment_id": payment_id}, {"_id": 0})
            if order and order.get("status") != "refunded":
                await db.orders.update_one({"order_id": order["order_id"]}, {"$set": {
                    "status": "refunded",
                    "pickup_code": None,
                    "pickup_verified": False,
                    "refunded_at": datetime.now(timezone.utc),
                    "refunded_by": "Razorpay (webhook)",
                }})
                logger.info(f"Webhook: order {order['order_id']} marked refunded ({event})")
                return {"status": "refunded", "order_id": order["order_id"]}
        return {"status": "refund_ignored"}

    # Extract the order id + payment id from either payment or order entity
    rzp_order_id = ""
    payment_id = ""
    try:
        if event in ("payment.captured", "payment.authorized"):
            entity = payload["payload"]["payment"]["entity"]
            rzp_order_id = entity.get("order_id", "")
            payment_id = entity.get("id", "")
        elif event == "order.paid":
            order_entity = payload["payload"]["order"]["entity"]
            rzp_order_id = order_entity.get("id", "")
            payment_id = payload.get("payload", {}).get("payment", {}).get("entity", {}).get("id", "")
    except Exception:
        pass

    if not rzp_order_id:
        return {"status": "ignored"}

    pending = await db.pending_orders.find_one({"razorpay_order_id": rzp_order_id}, {"_id": 0})
    if pending:
        order_id = await _finalize_order(pending, payment_id)
        logger.info(f"Webhook finalized order {order_id} for {rzp_order_id} ({event})")
        return {"status": "processed", "order_id": order_id}
    return {"status": "already_processed_or_unknown"}



@api.get("/orders/user")
async def user_orders(request: Request):
    user = await get_current_user(request)
    # Auto-expire stale reserved orders (pickup_end_time passed)
    now = datetime.now(timezone.utc)
    reserved = await db.orders.find({"user_id": user["user_id"], "status": "reserved"}, {"_id": 0}).to_list(500)
    for o in reserved:
        try:
            end_h, end_m = map(int, o.get("pickup_end_time", "23:59").split(":"))
            order_date = o.get("created_at", now)
            if hasattr(order_date, 'date'):
                expire_at = order_date.replace(hour=end_h, minute=end_m, second=0, tzinfo=timezone.utc)
                if expire_at < order_date:
                    expire_at += timedelta(days=1)
                if now > expire_at:
                    await db.orders.update_one({"order_id": o["order_id"]}, {"$set": {"status": "expired"}})
        except Exception:
            pass
    orders = await db.orders.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for o in orders:
        if "created_at" in o and hasattr(o["created_at"], "isoformat"):
            o["created_at"] = o["created_at"].isoformat()
    return orders

@api.put("/orders/{order_id}/cancel")
async def cancel_user_order(order_id: str, request: Request):
    user = await get_current_user(request)
    order = await db.orders.find_one({"order_id": order_id, "user_id": user["user_id"]}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order["status"] != "reserved":
        raise HTTPException(status_code=400, detail="Only reserved orders can be cancelled")
    await db.orders.update_one({"order_id": order_id}, {"$set": {"status": "cancelled"}})
    # Restore item quantity (if tracked)
    if order.get("quantity"):
        await db.menu_items.update_one(
            {"menu_item_id": order.get("food_item_id"), "quantity_available": {"$ne": None}},
            {"$inc": {"quantity_available": order.get("quantity", 0)}},
        )
    return {"message": "Order cancelled"}

# ══════════════════════════════════════════════════════════════════════════
#  HELP & SUPPORT
# ══════════════════════════════════════════════════════════════════════════

# Issue types config — add new entries here; the app renders from this too.
SUPPORT_ISSUE_TYPES = {
    "refund", "order_cancelled", "restaurant_closed", "wrong_item",
    "payment_issue", "pickup_expired", "app_bug", "other",
}
SUPPORT_ISSUE_LABELS = {
    "refund": "Refund", "order_cancelled": "Order Cancelled", "restaurant_closed": "Restaurant Closed",
    "wrong_item": "Wrong Item Received", "payment_issue": "Payment Issue", "pickup_expired": "Pickup Expired",
    "app_bug": "App Bug", "other": "Other",
}


def _smtp_configured() -> bool:
    host = os.environ.get("SMTP_HOST", "")
    user = os.environ.get("SMTP_USER", "")
    pwd = os.environ.get("SMTP_PASSWORD", "")
    # Treat placeholder values as "not configured"
    if not host or not user or not pwd:
        return False
    if "your-" in user.lower() or "your-" in pwd.lower() or "placeholder" in pwd.lower():
        return False
    return True


def _send_email_sync(subject: str, html_body: str, text_body: str) -> bool:
    if not _smtp_configured():
        logger.info(f"[support-email] SMTP not configured — skipping send. Subject: {subject}")
        return False
    host = os.environ["SMTP_HOST"]
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ["SMTP_USER"]
    pwd = os.environ["SMTP_PASSWORD"]
    sender = os.environ.get("SUPPORT_EMAIL_FROM", user)
    to_addr = os.environ.get("SUPPORT_EMAIL_TO", "anubhavg@perfectlygood.in")
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_addr
    msg.attach(MIMEText(text_body, "plain"))
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP(host, port, timeout=20) as server:
            server.starttls()
            server.login(user, pwd)
            server.sendmail(sender, [to_addr], msg.as_string())
        logger.info(f"[support-email] Sent to {to_addr}: {subject}")
        return True
    except Exception as e:
        logger.error(f"[support-email] Failed to send: {e}")
        return False


async def _send_support_email(subject: str, html_body: str, text_body: str) -> bool:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _send_email_sync, subject, html_body, text_body)


async def _recent_order_today(user_id: str) -> Optional[dict]:
    now = datetime.now(timezone.utc)
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    return await db.orders.find_one(
        {"user_id": user_id, "created_at": {"$gte": start}},
        {"_id": 0}, sort=[("created_at", -1)],
    )


@api.get("/support/context")
async def support_context(request: Request):
    """Auto-populate fields from the customer's most recent order placed today."""
    user = await get_current_user(request)
    order = await _recent_order_today(user["user_id"])
    ctx = {
        "customer_name": user.get("name", ""),
        "phone": user.get("phone", ""),
        "order_id": None, "restaurant_name": None, "order_amount": None,
        "pickup_datetime": None, "has_order": False,
    }
    if order:
        pickup = ""
        if order.get("pickup_start_time") or order.get("pickup_end_time"):
            pickup = f"{order.get('pickup_start_time','')} - {order.get('pickup_end_time','')}".strip(" -")
        created = order.get("created_at")
        date_str = created.strftime("%d %b %Y") if hasattr(created, "strftime") else ""
        ctx.update({
            "order_id": order.get("order_id"),
            "restaurant_name": order.get("vendor_name", ""),
            "order_amount": order.get("total_amount", 0),
            "pickup_datetime": f"{date_str}  {pickup}".strip(),
            "has_order": True,
        })
    return ctx


class SupportRequestBody(BaseModel):
    issue_type: str
    message: Optional[str] = ""
    photo_base64: Optional[str] = None
    device_model: Optional[str] = None
    app_version: Optional[str] = None
    what_happened: Optional[str] = None


@api.post("/support/requests")
async def create_support_request(body: SupportRequestBody, request: Request):
    user = await get_current_user(request)
    if body.issue_type not in SUPPORT_ISSUE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid issue type")
    if body.issue_type == "wrong_item" and not (body.photo_base64 or "").strip():
        raise HTTPException(status_code=400, detail="A photo is required for a wrong item report.")

    order = await _recent_order_today(user["user_id"])
    order_snapshot = {}
    if order:
        order_snapshot = {
            "order_id": order.get("order_id"),
            "restaurant_name": order.get("vendor_name", ""),
            "order_amount": order.get("total_amount", 0),
            "pickup_start_time": order.get("pickup_start_time", ""),
            "pickup_end_time": order.get("pickup_end_time", ""),
        }

    req_id = gen_id("support")
    now = datetime.now(timezone.utc)
    doc = {
        "support_id": req_id,
        "user_id": user["user_id"],
        "customer_name": user.get("name", ""),
        "phone": user.get("phone", ""),
        "email": user.get("email", ""),
        "issue_type": body.issue_type,
        "issue_label": SUPPORT_ISSUE_LABELS.get(body.issue_type, body.issue_type),
        "message": (body.message or "").strip(),
        "photo_base64": body.photo_base64 if body.issue_type == "wrong_item" else None,
        "device_model": body.device_model if body.issue_type == "app_bug" else None,
        "app_version": body.app_version if body.issue_type == "app_bug" else None,
        "what_happened": body.what_happened if body.issue_type == "app_bug" else None,
        "order": order_snapshot,
        "status": "open",
        "whatsapp_enabled": False,
        "whatsapp_enabled_by": None,
        "whatsapp_enabled_at": None,
        "resolved_by": None,
        "resolved_at": None,
        "created_at": now,
    }
    await db.support_requests.insert_one(doc)

    # Build + send email notification (non-blocking; best-effort)
    label = doc["issue_label"]
    rows = [
        ("Issue", label), ("Customer", doc["customer_name"]), ("Phone", doc["phone"]),
        ("Email", doc["email"]), ("Order ID", order_snapshot.get("order_id", "—")),
        ("Restaurant", order_snapshot.get("restaurant_name", "—")),
        ("Order Amount", f"₹{order_snapshot.get('order_amount', '—')}"),
        ("Pickup", f"{order_snapshot.get('pickup_start_time','')} - {order_snapshot.get('pickup_end_time','')}"),
        ("Message", doc["message"] or "—"),
    ]
    if body.issue_type == "app_bug":
        rows += [("Device", doc["device_model"] or "—"), ("App Version", doc["app_version"] or "—"),
                 ("What happened", doc["what_happened"] or "—")]
    if body.issue_type == "wrong_item":
        rows += [("Photo", "Attached (stored in dashboard)")]
    html = "<h2>New Support Request</h2><table style='border-collapse:collapse'>" + "".join(
        f"<tr><td style='padding:4px 12px;font-weight:bold'>{k}</td><td style='padding:4px 12px'>{v}</td></tr>" for k, v in rows
    ) + f"</table><p style='color:#888'>Request ID: {req_id}</p>"
    text = "New Support Request\n" + "\n".join(f"{k}: {v}" for k, v in rows) + f"\nRequest ID: {req_id}"
    email_sent = await _send_support_email(f"[Support] {label} — {doc['customer_name']}", html, text)

    return {"support_id": req_id, "message": "Support request submitted", "email_sent": email_sent}


def _support_public(d: dict) -> dict:
    """Shape a support ticket for the customer's own view (includes whatsapp state)."""
    return {
        "support_id": d.get("support_id"),
        "issue_type": d.get("issue_type"),
        "issue_label": d.get("issue_label"),
        "message": d.get("message", ""),
        "status": d.get("status", "open"),
        "whatsapp_enabled": bool(d.get("whatsapp_enabled")),
        "customer_name": d.get("customer_name", ""),
        "phone": d.get("phone", ""),
        "device_model": d.get("device_model"),
        "app_version": d.get("app_version"),
        "what_happened": d.get("what_happened"),
        "order": d.get("order", {}),
        "created_at": d.get("created_at").isoformat() if hasattr(d.get("created_at"), "isoformat") else d.get("created_at"),
    }


@api.get("/support/my-requests")
async def my_support_requests(request: Request):
    user = await get_current_user(request)
    rows = await db.support_requests.find(
        {"user_id": user["user_id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return [_support_public(r) for r in rows]


@api.get("/ops/support-open-count")
async def ops_support_open_count(request: Request):
    await require_permission(request, "manage_support")
    n = await db.support_requests.count_documents({"status": "open"})
    return {"open": n}


@api.get("/ops/support-requests")
async def ops_list_support(request: Request, issue_type: Optional[str] = None, status: Optional[str] = None,
                           page: int = 1, page_size: int = 50):
    await require_permission(request, "manage_support")
    query: dict = {}
    if issue_type:
        query["issue_type"] = issue_type
    if status:
        query["status"] = status
    total = await db.support_requests.count_documents(query)
    skip = max(page - 1, 0) * page_size
    rows = await db.support_requests.find(
        query, {"_id": 0, "photo_base64": 0}
    ).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for r in rows:
        if hasattr(r.get("created_at"), "isoformat"):
            r["created_at"] = r["created_at"].isoformat()
        r["order_id"] = (r.get("order") or {}).get("order_id")
        r["restaurant_name"] = (r.get("order") or {}).get("restaurant_name")
    return {"items": rows, "total": total, "page": page, "page_size": page_size}


@api.get("/ops/support-requests/{support_id}")
async def ops_support_detail(support_id: str, request: Request):
    await require_permission(request, "manage_support")
    d = await db.support_requests.find_one({"support_id": support_id}, {"_id": 0})
    if not d:
        raise HTTPException(status_code=404, detail="Support request not found")
    for k in ("created_at", "whatsapp_enabled_at", "resolved_at"):
        if hasattr(d.get(k), "isoformat"):
            d[k] = d[k].isoformat()
    return d


@api.put("/ops/support-requests/{support_id}/resolve")
async def ops_support_resolve(support_id: str, request: Request):
    user = await require_permission(request, "manage_support")
    result = await db.support_requests.update_one(
        {"support_id": support_id},
        {"$set": {"status": "resolved", "resolved_by": user.get("name") or user.get("email"),
                  "resolved_at": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Support request not found")
    return {"message": "Marked as resolved", "status": "resolved"}


@api.put("/ops/support-requests/{support_id}/whatsapp")
async def ops_support_enable_whatsapp(support_id: str, request: Request):
    user = await require_permission(request, "manage_support")
    result = await db.support_requests.update_one(
        {"support_id": support_id},
        {"$set": {"whatsapp_enabled": True, "whatsapp_enabled_by": user.get("name") or user.get("email"),
                  "whatsapp_enabled_at": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Support request not found")
    return {"message": "WhatsApp enabled", "whatsapp_enabled": True}




class CreateDropBody(BaseModel):
    menu_item_id: str
    discounted_price: float
    quantity_available: int
    pickup_start_time: str
    pickup_end_time: str
    expiry: Optional[str] = None

class ToggleDropBody(BaseModel):
    is_active: bool

class UpdateOrderStatusBody(BaseModel):
    status: str

class UpdateVendorProfileBody(BaseModel):
    address: Optional[str] = None
    phone: Optional[str] = None

@api.get("/vendor/profile")
async def vendor_profile(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    return vendor

@api.put("/vendor/profile")
async def update_vendor_profile(body: UpdateVendorProfileBody, request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    updates = {}
    if body.phone is not None:
        updates["phone"] = body.phone
    if body.address:
        location = geocode_address(body.address)
        if location:
            updates["location"] = location
        else:
            updates["location"] = {"lat": 0, "lon": 0, "address": body.address, "maps_url": f"https://www.google.com/maps/search/?api=1&query={body.address.replace(' ', '+')}"}
    if updates:
        await db.vendors.update_one({"vendor_id": vendor["vendor_id"]}, {"$set": updates})
    updated = await db.vendors.find_one({"vendor_id": vendor["vendor_id"]}, {"_id": 0})
    return updated

@api.get("/vendor/menu")
async def vendor_menu(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    items = await db.menu_items.find({"vendor_id": vendor["vendor_id"]}, {"_id": 0}).to_list(200)
    return items

@api.get("/vendor/drops")
async def vendor_drops(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    items = await db.menu_items.find({"vendor_id": vendor["vendor_id"]}, {"_id": 0}).to_list(500)
    return [item_to_drop(i, vendor) for i in items]

@api.post("/vendor/drops")
async def create_vendor_drop(body: CreateDropBody, request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    _require_staff_perm(user, "add_drops")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    menu_item = await db.menu_items.find_one({"menu_item_id": body.menu_item_id, "vendor_id": vendor["vendor_id"]}, {"_id": 0})
    if not menu_item:
        raise HTTPException(status_code=404, detail="Menu item not found")

    if vendor.get("status") != "active":
        raise HTTPException(status_code=403, detail="Your account must be approved before you can list surplus deals.")

    # Surplus listings must be at least 30% off the regular menu (dine-in) price
    op = menu_item.get("original_price") or 0
    if op > 0 and body.discounted_price > round(op * 0.7, 2):
        raise HTTPException(
            status_code=400,
            detail=f"Surplus price must be at least 30% off the menu price of ₹{op:.0f} (₹{round(op * 0.7):.0f} or less).",
        )

    updates = {
        "discounted_price": body.discounted_price,
        "quantity_available": body.quantity_available,
        "expiry": (body.expiry or "").strip(),
        "available_today": True,
        "updated_at": datetime.now(timezone.utc),
    }
    await db.menu_items.update_one({"menu_item_id": body.menu_item_id}, {"$set": updates})
    updated = await db.menu_items.find_one({"menu_item_id": body.menu_item_id}, {"_id": 0})

    # Notify all users about the new listing
    op = menu_item.get("original_price") or 0
    discount = round(((op - body.discounted_price) / op) * 100) if op else 0
    await send_push_to_all_users(
        title=f"{vendor.get('name', 'A vendor')} just dropped!",
        body=f"{menu_item['name']} — ₹{body.discounted_price} ({discount}% off)",
        data={"item_id": body.menu_item_id, "type": "new_drop"},
    )

    return item_to_drop(updated, vendor)

@api.put("/vendor/drops/{item_id}")
async def toggle_vendor_drop(item_id: str, body: ToggleDropBody, request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    _require_staff_perm(user, "add_drops")
    if body.is_active:
        vendor = await _resolve_vendor(user)
        if not vendor or vendor.get("status") != "active":
            raise HTTPException(status_code=403, detail="Your account must be approved before you can go live.")
    result = await db.menu_items.update_one({"menu_item_id": item_id}, {"$set": {"available_today": body.is_active, "updated_at": datetime.now(timezone.utc)}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Updated", "is_active": body.is_active}


class ToggleStockBody(BaseModel):
    in_stock: bool


class VendorMenuEditBody(BaseModel):
    image_url: Optional[str] = None
    kcal: Optional[int] = None
    protein: Optional[float] = None


@api.put("/vendor/menu/{item_id}")
async def vendor_edit_menu_item(item_id: str, body: VendorMenuEditBody, request: Request):
    """Vendor edits ONLY the image, kcal and protein of their own menu item.
    Name / price / description remain Ops-controlled."""
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    _require_staff_perm(user, "edit_menu")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    updates: dict = {"updated_at": datetime.now(timezone.utc)}
    if body.image_url is not None:
        updates["image_url"] = body.image_url
    if body.kcal is not None:
        updates["kcal"] = body.kcal
    if body.protein is not None:
        updates["protein"] = body.protein
    result = await db.menu_items.update_one(
        {"menu_item_id": item_id, "vendor_id": vendor["vendor_id"]},
        {"$set": updates},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    updated = await db.menu_items.find_one({"menu_item_id": item_id}, {"_id": 0})
    return updated


_IMG_MIME = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".webp": "image/webp", ".heic": "image/heic", ".heif": "image/heif", ".gif": "image/gif",
}
_MAX_IMG_BYTES = 5 * 1024 * 1024  # 5 MB per image


@api.post("/ops/vendors/{vendor_id}/bulk-images")
async def ops_bulk_upload_images(vendor_id: str, request: Request, file: UploadFile = File(...)):
    """Ops/Admin upload a ZIP of images for a specific vendor; each filename (minus
    extension) is matched case-insensitively to an existing menu item's name and only
    the image is updated. Never creates items or touches any other field."""
    user = await require_permission(request, "manage_menu")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if user.get("role") == "operations" and vendor.get("assigned_ops") != user["user_id"]:
        raise HTTPException(status_code=403, detail="This vendor is not assigned to you")
    raw = await file.read()
    return await _bulk_update_images(vendor, raw)


async def _bulk_update_images(vendor: dict, raw: bytes) -> dict:
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="That file is not a valid ZIP archive.")

    # Build a case-insensitive name -> item lookup for this vendor only.
    items = await db.menu_items.find({"vendor_id": vendor["vendor_id"]}, {"_id": 0, "menu_item_id": 1, "name": 1}).to_list(2000)
    by_name = {(i.get("name") or "").strip().lower(): i for i in items}

    matched, skipped = [], []
    seen_items: set = set()
    for info in zf.infolist():
        if info.is_dir():
            continue
        fname = os.path.basename(info.filename)
        if not fname or fname.startswith(".") or "__MACOSX" in info.filename:
            continue
        base, ext = os.path.splitext(fname)
        ext = ext.lower()
        if ext not in _IMG_MIME:
            skipped.append({"filename": fname, "reason": "Not an image file"})
            continue
        key = base.strip().lower()
        item = by_name.get(key)
        if not item:
            skipped.append({"filename": fname, "reason": "No matching menu item"})
            logger.info(f"[bulk-images] skip '{fname}' — no matching item for vendor {vendor['vendor_id']}")
            continue
        data = zf.read(info)
        if len(data) > _MAX_IMG_BYTES:
            skipped.append({"filename": fname, "reason": "Image larger than 5 MB"})
            continue
        b64 = f"data:{_IMG_MIME[ext]};base64,{base64.b64encode(data).decode()}"
        await db.menu_items.update_one(
            {"menu_item_id": item["menu_item_id"], "vendor_id": vendor["vendor_id"]},
            {"$set": {"image_url": b64, "updated_at": datetime.now(timezone.utc)}},
        )
        matched.append({"filename": fname, "item_name": item["name"]})
        seen_items.add(item["menu_item_id"])

    return {
        "updated_count": len(seen_items),
        "matched": matched,
        "skipped": skipped,
        "total_images": len(matched) + len(skipped),
    }


@api.put("/vendor/menu/{item_id}/toggle")
async def toggle_menu_stock(item_id: str, body: ToggleStockBody, request: Request):
    """Vendor toggles a regular menu item's Sold Out state.
    in_stock=False => sold out (hidden from customers + not orderable). We stamp
    the IST date it was marked so the daily midnight reset (and restart catch-up)
    can bring it back automatically. in_stock=True => available again."""
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    set_fields = {"in_stock": body.in_stock, "updated_at": datetime.now(timezone.utc)}
    set_fields["sold_out_at"] = None if body.in_stock else today_ist_str()
    result = await db.menu_items.update_one(
        {"menu_item_id": item_id, "vendor_id": vendor["vendor_id"]},
        {"$set": set_fields},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Updated", "in_stock": body.in_stock}

@api.get("/vendor/orders")
async def vendor_orders(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    orders = await db.orders.find({"vendor_id": vendor["vendor_id"]}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for o in orders:
        o["customer_name"] = o.get("user_name", "Customer")
        # Never expose the pickup code to the vendor — they must ask the customer to show it.
        o.pop("pickup_code", None)
        if "created_at" in o and hasattr(o["created_at"], "isoformat"):
            o["created_at"] = o["created_at"].isoformat()
    return orders


class VerifyPickupBody(BaseModel):
    code: str


@api.put("/vendor/orders/{order_id}/verify-pickup")
async def verify_pickup(order_id: str, body: VerifyPickupBody, request: Request):
    """Vendor verifies the customer's pickup code to complete the order.
    Completion only happens on a correct code; idempotent & race-safe."""
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    _require_staff_perm(user, "complete_orders")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    order = await db.orders.find_one({"order_id": order_id, "vendor_id": vendor["vendor_id"]}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") == "picked_up":
        raise HTTPException(status_code=400, detail="This order has already been completed.")
    if order.get("status") in ("cancelled", "refunded", "expired"):
        raise HTTPException(status_code=400, detail="This order can no longer be verified.")

    code = (body.code or "").strip()
    now = datetime.now(timezone.utc)
    # Atomic: only a reserved order with the matching code is completed — prevents
    # duplicate verification and wrong-code completion in one shot.
    result = await db.orders.update_one(
        {"order_id": order_id, "vendor_id": vendor["vendor_id"], "status": "reserved", "pickup_code": code},
        {"$set": {
            "status": "picked_up",
            "pickup_verified": True,
            "pickup_verified_at": now,
            "pickup_verified_by": user.get("name") or user.get("email") or "Vendor",
        }},
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Incorrect pickup code. Please ask the customer to show the correct code.")
    return {"message": "Pickup verified", "status": "picked_up", "order_id": order_id}

@api.put("/vendor/orders/{order_id}/status")
async def update_vendor_order_status(order_id: str, body: UpdateOrderStatusBody, request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    if body.status != "cancelled":
        raise HTTPException(status_code=400, detail="Orders are completed via pickup code verification.")
    result = await db.orders.update_one({"order_id": order_id, "status": {"$nin": ["picked_up", "refunded"]}}, {"$set": {"status": body.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found or cannot be cancelled")
    return {"message": "Status updated", "status": body.status}

# ── Vendor Payouts ──────────────────────────────────────────────────────

COMMISSION_RATE = 0.15
GST_ON_COMMISSION = 0.18


def order_revenue(o: dict) -> float:
    """Pre-tax item revenue for an order (price at time of sale)."""
    if o.get("item_subtotal") is not None:
        return round(o["item_subtotal"], 2)
    if o.get("discounted_price") is not None:
        return round(o["discounted_price"] * o.get("quantity", 1), 2)
    return round(o.get("total_amount", 0) / 1.10, 2)

@api.get("/vendor/payouts/summary")
async def vendor_payouts_summary(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    vid = vendor["vendor_id"]
    completed = await db.orders.find({"vendor_id": vid, "status": "picked_up"}, {"_id": 0}).to_list(10000)
    cfg = await get_settings_doc()
    total_revenue = round(sum(order_revenue(o) for o in completed), 2)
    total_commission = round(total_revenue * cfg["commission_rate"], 2)
    gst_on_commission = round(total_commission * cfg["gst_on_commission"], 2)
    total_deductions = round(total_commission + gst_on_commission, 2)
    net_earnings = round(total_revenue - total_deductions, 2)
    payouts = await db.payouts.find({"vendor_id": vid}, {"_id": 0}).to_list(10000)
    total_paid = round(sum(p.get("amount", 0) for p in payouts), 2)
    pending_payout = round(net_earnings - total_paid, 2)
    return {
        "total_orders_completed": len(completed),
        "total_revenue": total_revenue,
        "total_commission": total_commission,
        "gst_on_commission": gst_on_commission,
        "total_deductions": total_deductions,
        "net_earnings": net_earnings,
        "total_paid": total_paid,
        "pending_payout": pending_payout,
    }

@api.get("/vendor/payouts/orders")
async def vendor_payouts_orders(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    completed = await db.orders.find(
        {"vendor_id": vendor["vendor_id"], "status": "picked_up"}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    cfg = await get_settings_doc()
    result = []
    for o in completed:
        qty = o.get("quantity", 1)
        line_total = order_revenue(o)
        dp = round(line_total / max(qty, 1), 2)
        commission = round(line_total * cfg["commission_rate"], 2)
        gst_on_comm = round(commission * cfg["gst_on_commission"], 2)
        total_deduction = round(commission + gst_on_comm, 2)
        result.append({
            "order_id": o["order_id"],
            "food_item_name": o.get("food_item_name", ""),
            "quantity": qty,
            "discounted_price": dp,
            "vendor_earning": round(line_total - total_deduction, 2),
            "commission": commission,
            "gst_on_commission": gst_on_comm,
            "created_at": o["created_at"].isoformat() if hasattr(o.get("created_at"), "isoformat") else str(o.get("created_at", "")),
        })
    return result

# ══════════════════════════════════════════════════════════════════════════
#  ADMIN
# ══════════════════════════════════════════════════════════════════════════

class CreateVendorBody(BaseModel):
    name: str
    category: str
    email: str
    password: str
    phone: Optional[str] = ""
    location: Optional[dict] = None
    place_id: Optional[str] = None
    logo_url: Optional[str] = None
    service_type: Optional[str] = "both"  # "dine_in", "takeaway", "both"

class AddMenuItemBody(BaseModel):
    name: str
    description: Optional[str] = ""
    original_price: float
    image_url: Optional[str] = ""

class AddPayoutBody(BaseModel):
    vendor_id: str
    amount: float
    note: Optional[str] = ""

@api.get("/admin/vendors")
async def admin_vendors(request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(200)
    return vendors

@api.get("/admin/geocode")
async def admin_geocode(address: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    location = geocode_address(address)
    if not location:
        raise HTTPException(status_code=400, detail="Could not find this address. Try being more specific.")
    return location

@api.post("/admin/vendors")
async def admin_create_vendor(body: CreateVendorBody, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    email = body.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")

    user_id = gen_id("user")
    vendor_id = gen_id("vendor")

    # Resolve location from address or use provided location
    if body.place_id:
        # place_id field now used as address text for geocoding
        location = geocode_address(body.place_id)
        if not location:
            # Fallback: store address as-is without coordinates
            location = {"lat": 0, "lon": 0, "address": body.place_id, "maps_url": f"https://www.google.com/maps/search/?api=1&query={body.place_id.replace(' ', '+')}"}
    elif body.location:
        location = body.location
    else:
        location = {"lat": 12.9716, "lon": 77.5946, "address": "Bangalore"}

    await db.users.insert_one({
        "user_id": user_id,
        "email": email,
        "name": body.name,
        "password_hash": hash_password(body.password),
        "role": "vendor",
        "picture": None,
        "location": location,
        "created_at": datetime.now(timezone.utc),
    })
    vendor_doc = {
        "vendor_id": vendor_id,
        "user_id": user_id,
        "name": body.name,
        "category": body.category,
        "email": email,
        "phone": body.phone or "",
        "location": location,
        "logo_url": body.logo_url or "",
        "service_type": body.service_type or "both",
    }
    await db.vendors.insert_one(vendor_doc)
    vendor_doc.pop("_id", None)
    return vendor_doc

@api.delete("/admin/vendors/{vendor_id}")
async def admin_delete_vendor(vendor_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    await db.vendors.delete_one({"vendor_id": vendor_id})
    await db.users.delete_one({"user_id": vendor.get("user_id")})
    await db.menu_items.delete_many({"vendor_id": vendor_id})
    await db.drops.delete_many({"vendor_id": vendor_id})
    return {"message": "Vendor deleted"}

@api.get("/admin/vendors/{vendor_id}/menu")
async def admin_vendor_menu(vendor_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    items = await db.menu_items.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(200)
    return items

@api.post("/admin/vendors/{vendor_id}/menu")
async def admin_add_menu_item(vendor_id: str, body: AddMenuItemBody, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    menu_item_id = gen_id("menu")
    doc = {
        "menu_item_id": menu_item_id,
        "vendor_id": vendor_id,
        "name": body.name,
        "description": body.description or "",
        "original_price": body.original_price,
        "image_url": body.image_url or "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=600",
    }
    await db.menu_items.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/admin/menu-items/{menu_item_id}")
async def admin_delete_menu_item(menu_item_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    result = await db.menu_items.delete_one({"menu_item_id": menu_item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Menu item deleted"}

# ── Admin Payouts ───────────────────────────────────────────────────────

@api.post("/admin/payouts/add")
async def admin_add_payout(body: AddPayoutBody, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendor = await db.vendors.find_one({"vendor_id": body.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    payout_doc = {
        "payout_id": gen_id("payout"),
        "vendor_id": body.vendor_id,
        "amount": round(body.amount, 2),
        "note": body.note or "",
        "created_at": datetime.now(timezone.utc),
    }
    await db.payouts.insert_one(payout_doc)
    payout_doc.pop("_id", None)
    if hasattr(payout_doc.get("created_at"), "isoformat"):
        payout_doc["created_at"] = payout_doc["created_at"].isoformat()
    return payout_doc

@api.get("/admin/payouts/vendors")
async def admin_payouts_vendors(request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(200)
    cfg = await get_settings_doc()
    all_orders = await db.orders.find({"status": "picked_up"}, {"_id": 0, "vendor_id": 1, "item_subtotal": 1, "discounted_price": 1, "quantity": 1, "total_amount": 1}).to_list(100000)
    # Group orders by vendor
    orders_by_vendor: dict = {}
    for o in all_orders:
        vid = o.get("vendor_id")
        if vid:
            orders_by_vendor.setdefault(vid, []).append(o)
    # Batch fetch all payouts
    all_payouts = await db.payouts.find({}, {"_id": 0, "vendor_id": 1, "amount": 1}).to_list(100000)
    payouts_by_vendor: dict = {}
    for p in all_payouts:
        vid = p.get("vendor_id")
        if vid:
            payouts_by_vendor.setdefault(vid, []).append(p)
    result = []
    for v in vendors:
        vid = v["vendor_id"]
        completed = orders_by_vendor.get(vid, [])
        total_revenue = round(sum(order_revenue(o) for o in completed), 2)
        commission = round(total_revenue * cfg["commission_rate"], 2)
        gst_on_comm = round(commission * cfg["gst_on_commission"], 2)
        total_deductions = round(commission + gst_on_comm, 2)
        net_earnings = round(total_revenue - total_deductions, 2)
        vendor_payouts = payouts_by_vendor.get(vid, [])
        total_paid = round(sum(p.get("amount", 0) for p in vendor_payouts), 2)
        result.append({
            "vendor_id": vid,
            "vendor_name": v.get("name", ""),
            "total_orders_completed": len(completed),
            "net_earnings": net_earnings,
            "total_paid": total_paid,
            "pending_payout": round(net_earnings - total_paid, 2),
        })
    return result

@api.get("/admin/payouts/{vendor_id}/history")
async def admin_payout_history(vendor_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    payouts = await db.payouts.find({"vendor_id": vendor_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for p in payouts:
        if hasattr(p.get("created_at"), "isoformat"):
            p["created_at"] = p["created_at"].isoformat()
    return payouts

@api.post("/admin/upload")
async def admin_upload(file: UploadFile = File(...), request: Request = None):
    if request:
        user = await get_current_user(request)
        if user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Admin only")
    contents = await file.read()
    filename = f"upload_{secrets.token_hex(8)}_{file.filename}"
    upload_path = ROOT_DIR / "uploads" / filename
    upload_path.parent.mkdir(parents=True, exist_ok=True)
    with open(upload_path, "wb") as f:
        f.write(contents)
    return {"url": f"/uploads/{filename}", "filename": filename}

# ══════════════════════════════════════════════════════════════════════════
#  OPS DASHBOARD (internal operations team)
# ══════════════════════════════════════════════════════════════════════════

class OpsVendorBody(BaseModel):
    name: str
    owner_name: Optional[str] = ""
    email: str
    password: Optional[str] = None
    phone: Optional[str] = ""
    restaurant_phone: Optional[str] = ""
    category: str = "Restaurant"
    full_address: Optional[str] = ""
    maps_link: Optional[str] = ""
    service_type: Optional[str] = "both"
    pickup_start_time: Optional[str] = "18:00"
    pickup_end_time: Optional[str] = "21:00"
    status: Optional[str] = "draft"
    assigned_ops: Optional[str] = ""
    discount_percentage: Optional[float] = 0
    storefront_image: Optional[str] = ""

class OpsMenuItemBody(BaseModel):
    name: str
    description: Optional[str] = ""
    original_price: float
    discounted_price: Optional[float] = None
    category: Optional[str] = ""
    serving_size: Optional[str] = ""
    food_type: Optional[str] = "veg"
    contains_egg: Optional[bool] = False
    available_today: Optional[bool] = False
    quantity_available: Optional[int] = None
    image_url: Optional[str] = ""

class VendorNoteBody(BaseModel):
    note: str

class MarkPaidBody(BaseModel):
    vendor_id: str
    amount: float
    reference_number: Optional[str] = ""
    notes: Optional[str] = ""
    method: Optional[str] = "bank_transfer"

class StaffBody(BaseModel):
    name: str
    email: str
    password: Optional[str] = None
    role: str
    permission_overrides: Optional[dict] = None

class AvailabilityBody(BaseModel):
    available_today: bool


def _month_start(now: datetime) -> datetime:
    return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

def _day_start(now: datetime) -> datetime:
    return now.replace(hour=0, minute=0, second=0, microsecond=0)


@api.get("/ops/dashboard/stats")
async def ops_dashboard_stats(request: Request, range: Optional[str] = None):
    await require_permission(request, "view_dashboard")
    cfg = await get_settings_doc()
    now = datetime.now(timezone.utc)
    day0, week0, month0 = _day_start(now), _day_start(now) - timedelta(days=7), _month_start(now)

    total_vendors = await db.vendors.count_documents({})
    active_vendors = await db.vendors.count_documents({"status": {"$ne": "inactive"}})
    pending_vendors = await db.vendors.count_documents({"status": "pending"})
    live_items = await db.menu_items.count_documents({"available_today": True})

    orders_today = await db.orders.count_documents({"created_at": {"$gte": day0}, "status": {"$nin": ["cancelled", "refunded"]}})
    orders_week = await db.orders.count_documents({"created_at": {"$gte": week0}, "status": {"$nin": ["cancelled", "refunded"]}})

    today_orders = await db.orders.find({"created_at": {"$gte": day0}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(100000)
    month_orders = await db.orders.find({"created_at": {"$gte": month0}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(100000)
    revenue_today = round(sum(order_revenue(o) for o in today_orders), 2)
    revenue_month = round(sum(order_revenue(o) for o in month_orders), 2)

    # Date-range scoped metrics (today / week / month) for the dashboard toggle.
    rng = range if range in ("today", "week", "month") else "today"
    range_start = {"today": day0, "week": week0, "month": month0}[rng]
    range_orders_list = await db.orders.find({"created_at": {"$gte": range_start}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(100000)
    range_orders = len(range_orders_list)
    range_revenue = round(sum(order_revenue(o) for o in range_orders_list), 2)
    range_commission = round(range_revenue * cfg["commission_rate"], 2)

    completed = await db.orders.find({"status": "picked_up"}, {"_id": 0}).to_list(100000)
    commission_earned = round(sum(order_revenue(o) for o in completed) * cfg["commission_rate"], 2)

    # pending payouts across all vendors
    net_total = 0.0
    by_vendor: dict = {}
    for o in completed:
        by_vendor.setdefault(o.get("vendor_id"), 0.0)
        by_vendor[o.get("vendor_id")] += order_revenue(o)
    paid_rows = await db.payouts.find({}, {"_id": 0, "vendor_id": 1, "amount": 1}).to_list(100000)
    paid_by_vendor: dict = {}
    for p in paid_rows:
        paid_by_vendor[p.get("vendor_id")] = paid_by_vendor.get(p.get("vendor_id"), 0.0) + p.get("amount", 0)
    pending_payouts = 0.0
    for vid, rev in by_vendor.items():
        commission = rev * cfg["commission_rate"]
        net = rev - commission - (commission * cfg["gst_on_commission"])
        pending_payouts += max(net - paid_by_vendor.get(vid, 0.0), 0)

    return {
        "total_vendors": total_vendors,
        "active_vendors": active_vendors,
        "pending_vendors": pending_vendors,
        "live_menu_items": live_items,
        "orders_today": orders_today,
        "orders_week": orders_week,
        "revenue_today": revenue_today,
        "revenue_month": revenue_month,
        "commission_earned": commission_earned,
        "pending_payouts": round(pending_payouts, 2),
        "range": rng,
        "range_orders": range_orders,
        "range_revenue": range_revenue,
        "range_commission": range_commission,
    }


async def _vendor_aggregates(vendor_ids: list) -> dict:
    """Return {vendor_id: {menu_count, order_count, revenue}}."""
    agg = {vid: {"menu_count": 0, "order_count": 0, "revenue": 0.0} for vid in vendor_ids}
    menu = await db.menu_items.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0, "vendor_id": 1}).to_list(100000)
    for m in menu:
        if m["vendor_id"] in agg:
            agg[m["vendor_id"]]["menu_count"] += 1
    orders = await db.orders.find({"vendor_id": {"$in": vendor_ids}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(100000)
    for o in orders:
        vid = o.get("vendor_id")
        if vid in agg:
            agg[vid]["order_count"] += 1
            agg[vid]["revenue"] += order_revenue(o)
    for vid in agg:
        agg[vid]["revenue"] = round(agg[vid]["revenue"], 2)
    return agg


@api.get("/ops/vendors")
async def ops_list_vendors(request: Request, search: Optional[str] = None, category: Optional[str] = None,
                           status: Optional[str] = None, page: int = 1, page_size: int = 25):
    user = await require_permission(request, "view_vendors")
    query: dict = {}
    # Operations staff can only ever see the vendors assigned to them.
    if user.get("role") == "operations":
        query["assigned_ops"] = user["user_id"]
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
        ]
    if category:
        query["category"] = category
    if status:
        query["status"] = status
    total = await db.vendors.count_documents(query)
    skip = max(page - 1, 0) * page_size
    vendors = await db.vendors.find(query, {"_id": 0}).sort("name", 1).skip(skip).limit(page_size).to_list(page_size)
    agg = await _vendor_aggregates([v["vendor_id"] for v in vendors])
    ops_names = await _ops_name_map()
    for v in vendors:
        a = agg.get(v["vendor_id"], {})
        v["menu_count"] = a.get("menu_count", 0)
        v["order_count"] = a.get("order_count", 0)
        v["revenue"] = a.get("revenue", 0)
        v["assigned_ops_name"] = ops_names.get(v.get("assigned_ops", ""), "")
        for k in ("created_at", "updated_at", "last_order_date"):
            if isinstance(v.get(k), datetime):
                v[k] = v[k].isoformat()
    return {"items": vendors, "total": total, "page": page, "page_size": page_size}


@api.post("/ops/vendors")
async def ops_create_vendor(body: OpsVendorBody, request: Request):
    user = await require_permission(request, "manage_vendors")
    email = body.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    location = {}
    if body.full_address:
        location = geocode_address(body.full_address) or {
            "lat": 0, "lon": 0, "address": body.full_address,
            "maps_url": body.maps_link or f"https://www.google.com/maps/search/?api=1&query={body.full_address.replace(' ', '+')}",
        }
    if body.maps_link:
        location["maps_url"] = body.maps_link
    user_id = gen_id("user")
    vendor_id = gen_id("vendor")
    await db.users.insert_one({
        "user_id": user_id, "email": email, "name": body.name,
        "phone": body.phone or "",
        "password_hash": hash_password(body.password or secrets.token_urlsafe(8)),
        "role": "vendor", "picture": None, "location": location,
        "created_at": datetime.now(timezone.utc),
    })
    now = datetime.now(timezone.utc)
    # Operations staff can only create vendors assigned to themselves; admin assigns freely.
    assigned = user["user_id"] if user.get("role") == "operations" else (body.assigned_ops or "")
    vendor_doc = {
        "vendor_id": vendor_id, "user_id": user_id, "name": body.name,
        "owner_name": body.owner_name or "", "category": body.category, "email": email,
        "phone": body.phone or "", "restaurant_phone": body.restaurant_phone or "",
        "full_address": body.full_address or "", "maps_link": body.maps_link or "",
        "location": location, "logo_url": "", "service_type": body.service_type or "both",
        "pickup_start_time": body.pickup_start_time or "18:00", "pickup_end_time": body.pickup_end_time or "21:00",
        "status": body.status or "draft", "assigned_ops": assigned,
        "discount_percentage": max(0.0, min(float(body.discount_percentage or 0), 90.0)),
        "storefront_image": body.storefront_image or "",
        "notes": [], "created_at": now, "updated_at": now, "last_order_date": None,
    }
    await db.vendors.insert_one(vendor_doc)
    vendor_doc.pop("_id", None)
    for k in ("created_at", "updated_at"):
        vendor_doc[k] = vendor_doc[k].isoformat()
    return vendor_doc


@api.get("/ops/vendors/{vendor_id}")
async def ops_vendor_detail(vendor_id: str, request: Request):
    user = await require_permission(request, "view_vendors")
    cfg = await get_settings_doc()
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if user.get("role") == "operations" and v.get("assigned_ops") != user["user_id"]:
        raise HTTPException(status_code=403, detail="This vendor is not assigned to you")
    ops_names = await _ops_name_map()
    v["assigned_ops_name"] = ops_names.get(v.get("assigned_ops", ""), "")
    menu = await db.menu_items.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(1000)
    for m in menu:
        for k in ("created_at", "updated_at"):
            if isinstance(m.get(k), datetime):
                m[k] = m[k].isoformat()
    orders = await db.orders.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(100000)
    completed = [o for o in orders if o.get("status") == "picked_up"]
    revenue = round(sum(order_revenue(o) for o in completed), 2)
    commission = round(revenue * cfg["commission_rate"], 2)
    payouts = await db.payouts.find({"vendor_id": vendor_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for p in payouts:
        if isinstance(p.get("created_at"), datetime):
            p["created_at"] = p["created_at"].isoformat()
    total_paid = round(sum(p.get("amount", 0) for p in payouts), 2)
    net = round(revenue - commission - (commission * cfg["gst_on_commission"]), 2)
    for k in ("created_at", "updated_at", "last_order_date"):
        if isinstance(v.get(k), datetime):
            v[k] = v[k].isoformat()
    v["menu_items"] = menu
    v["total_orders"] = len([o for o in orders if o.get("status") not in ("cancelled", "refunded")])
    v["completed_orders"] = len(completed)
    v["revenue"] = revenue
    v["commission"] = commission
    v["net_payable"] = net
    v["total_paid"] = total_paid
    v["pending_payout"] = round(net - total_paid, 2)
    v["payout_history"] = payouts
    return v


@api.put("/ops/vendors/{vendor_id}")
async def ops_update_vendor(vendor_id: str, body: OpsVendorBody, request: Request):
    user = await require_permission(request, "manage_vendors")
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if user.get("role") == "operations" and v.get("assigned_ops") != user["user_id"]:
        raise HTTPException(status_code=403, detail="This vendor is not assigned to you")
    updates = {
        "name": body.name, "owner_name": body.owner_name or "", "category": body.category,
        "phone": body.phone or "", "restaurant_phone": body.restaurant_phone or "",
        "full_address": body.full_address or "", "maps_link": body.maps_link or "",
        "service_type": body.service_type or "both", "pickup_start_time": body.pickup_start_time or "18:00",
        "pickup_end_time": body.pickup_end_time or "21:00", "status": body.status or "active",
        "discount_percentage": max(0.0, min(float(body.discount_percentage or 0), 90.0)),
        "storefront_image": body.storefront_image or "",
        "updated_at": datetime.now(timezone.utc),
    }
    # Only admins may (re)assign vendors to an ops member.
    if user.get("role") == "admin":
        updates["assigned_ops"] = body.assigned_ops or ""
    if body.full_address and body.full_address != v.get("full_address"):
        loc = geocode_address(body.full_address)
        if loc:
            if body.maps_link:
                loc["maps_url"] = body.maps_link
            updates["location"] = loc
    await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": updates})
    if body.name or body.phone:
        await db.users.update_one({"user_id": v.get("user_id")}, {"$set": {"name": body.name, "phone": body.phone or ""}})
    return {"message": "Vendor updated"}


@api.put("/ops/vendors/{vendor_id}/status")
async def ops_vendor_status(vendor_id: str, body: dict, request: Request):
    user = await require_permission(request, "manage_vendors")
    status = body.get("status", "active")
    if status == "active" and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can activate a vendor.")
    result = await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc)}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Status updated", "status": status}


# ══════════════════════════════════════════════════════════════════════════
#  VENDOR STAFF MANAGEMENT (a vendor manages its own staff accounts)
# ══════════════════════════════════════════════════════════════════════════

VENDOR_STAFF_PERMS = ["add_drops", "complete_orders", "edit_menu"]

def _staff_view(u: dict) -> dict:
    return {
        "user_id": u.get("user_id"),
        "name": u.get("name", ""),
        "email": u.get("email", ""),
        "permissions": u.get("staff_permissions") or [],
        "created_at": u.get("created_at").isoformat() if hasattr(u.get("created_at"), "isoformat") else u.get("created_at"),
    }

async def _require_vendor_owner(request: Request):
    """Only a vendor OWNER (role 'vendor') may manage staff — not staff themselves."""
    user = await get_current_user(request)
    if user.get("role") != "vendor":
        raise HTTPException(status_code=403, detail="Only the vendor owner can manage staff")
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    return user, vendor

def _clean_perms(perms):
    return [p for p in (perms or []) if p in VENDOR_STAFF_PERMS]

@api.get("/vendor/staff")
async def list_vendor_staff(request: Request):
    _, vendor = await _require_vendor_owner(request)
    rows = await db.users.find({"role": "vendor_staff", "parent_vendor_id": vendor["vendor_id"]}, {"_id": 0}).to_list(200)
    return {"items": [_staff_view(u) for u in rows], "permissions_available": VENDOR_STAFF_PERMS}

@api.post("/vendor/staff")
async def create_vendor_staff(body: dict, request: Request):
    _, vendor = await _require_vendor_owner(request)
    name = (body.get("name") or "").strip()
    email = (body.get("email") or "").strip().lower()
    password = (body.get("password") or "").strip()
    perms = _clean_perms(body.get("permissions"))
    if not name:
        raise HTTPException(status_code=400, detail="Staff name is required")
    if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
        raise HTTPException(status_code=400, detail="Please enter a valid email address")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="That email is already in use")
    now = datetime.now(timezone.utc)
    doc = {
        "user_id": gen_id("user"), "email": email, "name": name,
        "password_hash": hash_password(password), "role": "vendor_staff",
        "parent_vendor_id": vendor["vendor_id"], "staff_permissions": perms,
        "created_at": now,
    }
    await db.users.insert_one(doc)
    return _staff_view(doc)

@api.put("/vendor/staff/{user_id}")
async def update_vendor_staff(user_id: str, body: dict, request: Request):
    _, vendor = await _require_vendor_owner(request)
    staff = await db.users.find_one({"user_id": user_id, "role": "vendor_staff", "parent_vendor_id": vendor["vendor_id"]}, {"_id": 0})
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    updates: dict = {}
    if "permissions" in body:
        updates["staff_permissions"] = _clean_perms(body.get("permissions"))
    if body.get("name"):
        updates["name"] = body["name"].strip()
    if body.get("password"):
        if len((body["password"] or "").strip()) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        updates["password_hash"] = hash_password(body["password"].strip())
    if updates:
        await db.users.update_one({"user_id": user_id}, {"$set": updates})
    fresh = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return _staff_view(fresh)

@api.delete("/vendor/staff/{user_id}")
async def delete_vendor_staff(user_id: str, request: Request):
    _, vendor = await _require_vendor_owner(request)
    res = await db.users.delete_one({"user_id": user_id, "role": "vendor_staff", "parent_vendor_id": vendor["vendor_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Staff member not found")
    return {"message": "Staff member removed"}



# ══════════════════════════════════════════════════════════════════════════
#  VENDOR COMPLIANCE, VERIFICATION & AGREEMENT
# ══════════════════════════════════════════════════════════════════════════

VENDOR_STATUSES = ["draft", "pending_verification", "active", "rejected", "suspended"]

_AGREEMENT_TEXT = """VENDOR AGREEMENT
Perfectly Good
Prajjval Ventures Private Limited
Office No. 1190/1, 4th Floor, HSR Layout, Sector 3, 22nd Cross Road. Bengaluru 560102

Effective Date: The date on which the Vendor electronically accepts this Agreement through the Perfectly Good Vendor Compliance & Agreement Form.

This Vendor Agreement (\u201cAgreement\u201d) is entered into between:
Prajjval Ventures Private Limited, operating the Perfectly Good surplus food platform, having its registered office at Office No. 1190/1, 4th Floor, HSR Layout, Sector 3, 22nd Cross Road. Bengaluru 560102 (\u201cPlatform\u201d, \u201cwe\u201d, \u201cus\u201d or \u201cour\u201d), and
Any food business operator that completes the Perfectly Good Vendor Compliance & Agreement Form and electronically accepts this Agreement (\u201cVendor\u201d, \u201cyou\u201d or \u201cyour\u201d).
The Platform and the Vendor are individually referred to as a \u201cParty\u201d and collectively as the \u201cParties\u201d.

1. Purpose
1.1 The Platform operates a marketplace that connects customers with food business operators offering surplus, unsold or excess food at discounted prices, available for dine-in, takeaway or both, depending on the format selected by the Vendor during onboarding.
1.2 This Agreement sets out the terms on which the Vendor may list, prepare, package, hand over or serve food through the Platform.

2. Vendor Obligations
2.1 The Vendor shall maintain a valid FSSAI licence and all licences, registrations and permits required under applicable laws to prepare and sell food in India.
2.2 The Vendor shall ensure that all food listed on the Platform is fit for human consumption and prepared, stored, handled and packaged in accordance with the Food Safety and Standards Act, 2006 and all applicable rules and regulations.
2.3 The Vendor shall accurately disclose preparation time, best before or use by details and known allergens for every listed item.
2.4 For dine-in orders, the Vendor shall maintain hygienic seating and service areas in compliance with applicable health and safety requirements.
2.5 For takeaway orders, the Vendor shall use suitable food-grade packaging to ensure the food remains safe until the disclosed consumption period.
2.6 The Vendor shall not list spoiled, expired, contaminated or otherwise unsafe food.
2.7 The Vendor shall cooperate with the Platform during investigations relating to customer complaints and shall provide any requested records, photographs or other relevant information within twenty-four (24) hours of request.

3. Liability for Food Quality and Safety
3.1 The Vendor is solely responsible for the quality, safety and fitness for consumption of all food listed, prepared, packaged or served through the Platform.
3.2 Where the Platform determines, after reviewing a complaint, that the issue originated from the food supplied by the Vendor, the Vendor shall bear full responsibility for any refund, replacement, compensation, regulatory action, legal claim or other liability arising from that issue.
3.3 The Platform\u2019s determination regarding responsibility under this Agreement shall be final for the purpose of allocating liability between the Parties.
3.4 The Platform operates solely as a technology marketplace connecting customers with Vendors. The Platform does not prepare, cook, package, store or inspect food before it reaches customers and does not guarantee the quality or safety of any food listed by Vendors.
3.5 The Platform shall not be liable for any illness, injury, loss, damage or claim arising from food supplied by the Vendor where the issue relates to the quality, safety or suitability of the food.

4. Indemnity
4.1 The Vendor agrees to indemnify and hold harmless the Platform, its directors, officers, employees and representatives against all claims, losses, damages, liabilities, costs and legal expenses arising from:
- unsafe or non-compliant food supplied by the Vendor;
- any breach of this Agreement by the Vendor;
- any claim by a customer, regulator or third party relating to food supplied by the Vendor.

5. Listings, Orders and Payments
5.1 The Vendor shall ensure that all listings accurately state pricing, quantity, availability and fulfilment format.
5.2 The Platform shall remit payments for completed and undisputed orders in accordance with the payout schedule and commission structure communicated during onboarding.
5.3 Any refunds or compensation payable due to Vendor-related issues may be deducted from future payouts or recovered separately if necessary.

6. Complaint Investigation
6.1 The Platform may investigate complaints using customer feedback, photographs, timestamps, order details and any other available evidence.
6.2 The Vendor shall be given a reasonable opportunity to provide its explanation and supporting evidence before a decision is made.
6.3 If the Platform determines that the complaint resulted from the Vendor\u2019s food, liability shall be allocated in accordance with this Agreement.

7. Term and Termination
7.1 This Agreement commences on the Effective Date and continues until terminated.
7.2 Either Party may terminate this Agreement by providing thirty (30) days\u2019 written notice.
7.3 The Platform may suspend or terminate the Vendor immediately if the Vendor repeatedly supplies unsafe food or no longer maintains the licences required by law.
7.4 Termination shall not affect any rights or liabilities that accrued before termination.

8. Representations and Warranties
The Vendor represents and warrants that:
- it is legally authorised to enter into this Agreement;
- it holds all licences required by law;
- all information submitted during onboarding is accurate and complete;
- all food supplied through the Platform complies with applicable laws and food safety standards.

9. Independent Contractor
Nothing in this Agreement creates a partnership, joint venture, employment or agency relationship. The Vendor remains an independent business and is solely responsible for its employees, operations and food preparation.

10. Confidentiality
Each Party shall keep confidential all non-public business, operational and commercial information received from the other Party and shall use such information only for purposes relating to this Agreement.

11. Governing Law
This Agreement shall be governed by the laws of India. The courts at Bengaluru, Karnataka, India shall have exclusive jurisdiction over any dispute arising out of or relating to this Agreement.

12. General
12.1 This Agreement constitutes the complete agreement between the Parties and supersedes all previous discussions relating to its subject matter.
12.2 Any amendment shall be effective only if made in writing by the Platform.
12.3 If any provision is held invalid or unenforceable, the remaining provisions shall remain in full force and effect.
12.4 The Vendor may not assign or transfer its rights or obligations under this Agreement without the prior written consent of the Platform.

Electronic Acceptance
By completing and submitting the Perfectly Good Vendor Compliance & Agreement Form, the Vendor confirms that:
- it has read and understood this Agreement in full;
- it is authorised to enter into this Agreement on behalf of the business;
- all information and documents submitted are true, accurate and complete;
- it agrees to be legally bound by the terms of this Agreement.
The Vendor acknowledges that typing its full legal name as an electronic signature and submitting the onboarding form constitutes a valid electronic acceptance of this Agreement. The submission timestamp recorded by Perfectly Good shall constitute the Effective Date of this Agreement.

For Perfectly Good
Company: Prajjval Ventures Private Limited
Brand: Perfectly Good
Authorised Signatory: Anubhav Giri
Designation: Founder"""


async def get_agreement_doc() -> dict:
    doc = await db.settings.find_one({"_id": "vendor_agreement"})
    if not doc:
        doc = {"_id": "vendor_agreement", "version": "1.0", "content": _AGREEMENT_TEXT,
               "pdf_url": "", "updated_at": datetime.now(timezone.utc), "updated_by": "System"}
        await db.settings.insert_one(doc)
    doc.pop("_id", None)
    if isinstance(doc.get("updated_at"), datetime):
        doc["updated_at"] = doc["updated_at"].isoformat()
    return doc


def _iso(dt):
    return dt.isoformat() if isinstance(dt, datetime) else dt


def _cert(c):
    if not c or not isinstance(c, dict):
        return None
    if not (c.get("data") or ""):
        return None
    return {"name": c.get("name", "document"), "mime": c.get("mime", ""), "data": c.get("data")}


class SaveVerificationBody(BaseModel):
    business_name: Optional[str] = ""
    authorised_representative: Optional[str] = ""
    business_email: Optional[str] = ""
    gst_status: Optional[str] = ""
    gst_number: Optional[str] = ""
    gst_certificate: Optional[dict] = None
    fssai_number: Optional[str] = ""
    fssai_certificate: Optional[dict] = None
    bank_account_holder: Optional[str] = ""
    bank_account_number: Optional[str] = ""
    bank_ifsc: Optional[str] = ""
    bank_name: Optional[str] = ""


class SubmitVerificationBody(SaveVerificationBody):
    agreement_version: Optional[str] = ""
    signature_full_name: Optional[str] = ""
    signature_designation: Optional[str] = ""
    agreed_agreement: Optional[bool] = False
    decl_authorised: Optional[bool] = False
    decl_accurate: Optional[bool] = False
    decl_agreement: Optional[bool] = False
    decl_food_safety: Optional[bool] = False


async def _get_vendor_for_user(request: Request):
    user = await get_current_user(request)
    if user["role"] not in VENDOR_ROLES:
        raise HTTPException(status_code=403, detail="Not a vendor")
    vendor = await _resolve_vendor(user)
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor profile not found")
    return user, vendor


def _apply_verification(existing: dict, body: SaveVerificationBody) -> dict:
    existing = existing or {}
    return {
        "business_name": body.business_name or "",
        "authorised_representative": body.authorised_representative or "",
        "business_email": (body.business_email or "").strip(),
        "gst_status": body.gst_status or "",
        "gst_number": (body.gst_number or "").strip(),
        "gst_certificate": _cert(body.gst_certificate) or existing.get("gst_certificate"),
        "fssai_number": (body.fssai_number or "").strip(),
        "fssai_certificate": _cert(body.fssai_certificate) or existing.get("fssai_certificate"),
        "bank_account_holder": body.bank_account_holder or "",
        "bank_account_number": (body.bank_account_number or "").strip(),
        "bank_ifsc": (body.bank_ifsc or "").strip().upper(),
        "bank_name": body.bank_name or "",
        "agreement": existing.get("agreement"),
        "declarations": existing.get("declarations"),
    }


@api.get("/vendor/verification")
async def get_vendor_verification(request: Request):
    user, vendor = await _get_vendor_for_user(request)
    return {
        "status": vendor.get("status", "draft"),
        "rejection_reason": vendor.get("rejection_reason", ""),
        "locked": vendor.get("status") in ("pending_verification", "active", "suspended"),
        "verification": vendor.get("verification") or {},
        "agreement": await get_agreement_doc(),
        "submitted_at": _iso(vendor.get("verification_submitted_at")),
    }


@api.get("/vendor/agreement")
async def vendor_get_agreement(request: Request):
    await get_current_user(request)
    return await get_agreement_doc()


@api.put("/vendor/verification")
async def save_vendor_verification(body: SaveVerificationBody, request: Request):
    user, vendor = await _get_vendor_for_user(request)
    if vendor.get("status") in ("pending_verification", "active", "suspended"):
        raise HTTPException(status_code=400, detail="Verification is locked and cannot be edited right now.")
    v = _apply_verification(vendor.get("verification") or {}, body)
    await db.vendors.update_one({"vendor_id": vendor["vendor_id"]}, {"$set": {"verification": v, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Saved", "verification": v}


def _validate_verification(v: dict, body: SubmitVerificationBody):
    errors = []
    if not (v.get("business_name") or "").strip(): errors.append("Business name")
    if not (v.get("authorised_representative") or "").strip(): errors.append("Authorised representative")
    if not (v.get("business_email") or "").strip(): errors.append("Business email")
    gst_status = v.get("gst_status")
    if gst_status not in ("registered", "not_registered"):
        errors.append("GST status")
    elif gst_status == "registered":
        if not (v.get("gst_number") or "").strip(): errors.append("GST number")
        if not v.get("gst_certificate"): errors.append("GST certificate")
    if not (v.get("fssai_number") or "").strip(): errors.append("FSSAI licence number")
    if not v.get("fssai_certificate"): errors.append("FSSAI certificate")
    if not (v.get("bank_account_holder") or "").strip(): errors.append("Account holder name")
    if not (v.get("bank_account_number") or "").strip(): errors.append("Account number")
    if not (v.get("bank_ifsc") or "").strip(): errors.append("IFSC code")
    if not (v.get("bank_name") or "").strip(): errors.append("Bank name")
    if not (body.signature_full_name or "").strip(): errors.append("Signature full legal name")
    if not (body.signature_designation or "").strip(): errors.append("Signature designation")
    if not (body.agreed_agreement and body.decl_authorised and body.decl_accurate and body.decl_agreement and body.decl_food_safety):
        errors.append("All agreement checkboxes and declarations")
    return errors


@api.post("/vendor/verification/submit")
async def submit_vendor_verification(body: SubmitVerificationBody, request: Request):
    user, vendor = await _get_vendor_for_user(request)
    if vendor.get("status") in ("pending_verification", "active", "suspended"):
        raise HTTPException(status_code=400, detail="Your verification is already submitted or your account is active.")
    v = _apply_verification(vendor.get("verification") or {}, body)
    errors = _validate_verification(v, body)
    if errors:
        raise HTTPException(status_code=400, detail="Please complete: " + ", ".join(errors))
    agreement = await get_agreement_doc()
    now = datetime.now(timezone.utc)
    v["agreement"] = {
        "version": agreement.get("version", ""),
        "accepted": True,
        "accepted_at": now.isoformat(),
        "signature_full_name": (body.signature_full_name or "").strip(),
        "signature_designation": (body.signature_designation or "").strip(),
    }
    v["declarations"] = {"authorised": True, "accurate": True, "agreement": True, "food_safety": True}
    await db.vendors.update_one({"vendor_id": vendor["vendor_id"]}, {"$set": {
        "verification": v, "status": "pending_verification",
        "verification_submitted_at": now, "rejection_reason": "", "updated_at": now,
    }})
    return {"message": "Your verification has been submitted and is awaiting admin approval.", "status": "pending_verification"}


# ── Admin: vendor agreement management ──────────────────────────────────

class AgreementBody(BaseModel):
    content: str
    version: Optional[str] = None
    pdf_url: Optional[str] = ""
    bump_version: Optional[bool] = True


@api.get("/ops/vendor-agreement")
async def ops_get_agreement(request: Request):
    await require_permission(request, "view_vendors")
    return await get_agreement_doc()


@api.put("/ops/vendor-agreement")
async def ops_update_agreement(body: AgreementBody, request: Request):
    user = await require_permission(request, "manage_vendors")
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can edit the vendor agreement.")
    current = await get_agreement_doc()
    version = body.version
    if not version:
        if body.bump_version:
            try:
                version = str(round(float(current.get("version", "1.0")) + 0.1, 1))
            except Exception:
                version = "1.0"
        else:
            version = current.get("version", "1.0")
    await db.settings.update_one({"_id": "vendor_agreement"}, {"$set": {
        "content": body.content, "version": version, "pdf_url": body.pdf_url or "",
        "updated_at": datetime.now(timezone.utc), "updated_by": user.get("name") or user.get("email"),
    }}, upsert=True)
    return await get_agreement_doc()


# ── Admin: compliance review ────────────────────────────────────────────

class ReasonBody(BaseModel):
    reason: Optional[str] = ""


def _admin_only(user):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can perform this action.")


@api.get("/ops/compliance")
async def ops_compliance_list(request: Request, status: Optional[str] = None):
    await require_permission(request, "view_vendors")
    query: dict = {"status": status} if status else {}
    vendors = await db.vendors.find(query, {"_id": 0}).to_list(1000)
    out = []
    for v in vendors:
        ver = v.get("verification") or {}
        out.append({
            "vendor_id": v.get("vendor_id"), "name": v.get("name"), "email": v.get("email"),
            "status": v.get("status", "draft"),
            "business_name": ver.get("business_name", ""),
            "fssai_number": ver.get("fssai_number", ""),
            "gst_status": ver.get("gst_status", ""),
            "submitted_at": _iso(v.get("verification_submitted_at")),
            "rejection_reason": v.get("rejection_reason", ""),
        })
    order = {"pending_verification": 0, "rejected": 1, "draft": 2, "suspended": 3, "active": 4}
    out.sort(key=lambda r: (order.get(r["status"], 9), (r.get("name") or "").lower()))
    return {"items": out, "total": len(out)}


@api.get("/ops/compliance/{vendor_id}")
async def ops_compliance_detail(vendor_id: str, request: Request):
    await require_permission(request, "view_vendors")
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {
        "vendor_id": v.get("vendor_id"), "name": v.get("name"), "email": v.get("email"),
        "status": v.get("status", "draft"), "rejection_reason": v.get("rejection_reason", ""),
        "submitted_at": _iso(v.get("verification_submitted_at")),
        "verification": v.get("verification") or {},
    }


@api.post("/ops/vendors/{vendor_id}/approve")
async def ops_approve_vendor(vendor_id: str, request: Request):
    user = await require_permission(request, "manage_vendors")
    _admin_only(user)
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    now = datetime.now(timezone.utc)
    await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": {
        "status": "active", "rejection_reason": "",
        "approved_by": user.get("name") or user.get("email"), "approved_at": now, "updated_at": now,
    }})
    return {"message": "Vendor approved and activated", "status": "active"}


@api.post("/ops/vendors/{vendor_id}/reject")
async def ops_reject_vendor(vendor_id: str, body: ReasonBody, request: Request):
    user = await require_permission(request, "manage_vendors")
    _admin_only(user)
    if not (body.reason or "").strip():
        raise HTTPException(status_code=400, detail="A rejection reason is required.")
    now = datetime.now(timezone.utc)
    result = await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": {
        "status": "rejected", "rejection_reason": body.reason.strip(),
        "rejected_by": user.get("name") or user.get("email"), "rejected_at": now, "updated_at": now,
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor rejected", "status": "rejected"}


@api.post("/ops/vendors/{vendor_id}/suspend")
async def ops_suspend_vendor(vendor_id: str, body: ReasonBody, request: Request):
    user = await require_permission(request, "manage_vendors")
    _admin_only(user)
    now = datetime.now(timezone.utc)
    result = await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": {
        "status": "suspended", "suspension_reason": (body.reason or "").strip(),
        "suspended_by": user.get("name") or user.get("email"), "suspended_at": now, "updated_at": now,
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor suspended", "status": "suspended"}


@api.delete("/ops/vendors/{vendor_id}")
async def ops_delete_vendor(vendor_id: str, request: Request):
    user = await require_permission(request, "manage_vendors")
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can delete vendors")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    await db.vendors.delete_one({"vendor_id": vendor_id})
    await db.users.delete_one({"user_id": vendor.get("user_id")})
    await db.menu_items.delete_many({"vendor_id": vendor_id})
    return {"message": "Vendor deleted"}


@api.post("/ops/vendors/{vendor_id}/notes")
async def ops_add_note(vendor_id: str, body: VendorNoteBody, request: Request):
    user = await require_permission(request, "add_notes")
    note = {"note": body.note, "by": user.get("name", "Staff"), "at": datetime.now(timezone.utc).isoformat()}
    result = await db.vendors.update_one({"vendor_id": vendor_id}, {"$push": {"notes": note}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return note


@api.get("/ops/vendors/{vendor_id}/menu")
async def ops_vendor_menu(vendor_id: str, request: Request):
    await require_permission(request, "view_vendors")
    items = await db.menu_items.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(1000)
    for m in items:
        for k in ("created_at", "updated_at"):
            if isinstance(m.get(k), datetime):
                m[k] = m[k].isoformat()
    return items


@api.post("/ops/vendors/{vendor_id}/menu")
async def ops_add_menu_item(vendor_id: str, body: OpsMenuItemBody, request: Request):
    await require_permission(request, "manage_menu")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    cfg = await get_settings_doc()
    dp = body.discounted_price
    if dp is None:
        dp = round(body.original_price * (1 - cfg["default_discount_pct"] / 100), 2)
    now = datetime.now(timezone.utc)
    doc = {
        "menu_item_id": gen_id("menu"), "vendor_id": vendor_id, "name": body.name,
        "description": body.description or "", "original_price": body.original_price,
        "discounted_price": dp, "category": body.category or vendor.get("category", ""),
        "serving_size": body.serving_size or "", "food_type": body.food_type or "veg",
        "contains_egg": bool(body.contains_egg), "available_today": bool(body.available_today),
        "in_stock": True, "quantity_available": body.quantity_available,
        "image_url": body.image_url or "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=600",
        "created_at": now, "updated_at": now,
    }
    await db.menu_items.insert_one(doc)
    doc.pop("_id", None)
    for k in ("created_at", "updated_at"):
        doc[k] = doc[k].isoformat()
    return doc


@api.put("/ops/menu/{menu_item_id}")
async def ops_update_menu_item(menu_item_id: str, body: OpsMenuItemBody, request: Request):
    await require_permission(request, "manage_menu")
    updates = {
        "name": body.name, "description": body.description or "", "original_price": body.original_price,
        "category": body.category or "", "serving_size": body.serving_size or "",
        "food_type": body.food_type or "veg", "contains_egg": bool(body.contains_egg),
        "available_today": bool(body.available_today), "updated_at": datetime.now(timezone.utc),
    }
    if body.discounted_price is not None:
        updates["discounted_price"] = body.discounted_price
    if body.quantity_available is not None:
        updates["quantity_available"] = body.quantity_available
    if body.image_url:
        updates["image_url"] = body.image_url
    result = await db.menu_items.update_one({"menu_item_id": menu_item_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Menu item updated"}


@api.post("/ops/menu/{menu_item_id}/duplicate")
async def ops_duplicate_menu_item(menu_item_id: str, request: Request):
    await require_permission(request, "manage_menu")
    item = await db.menu_items.find_one({"menu_item_id": menu_item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Menu item not found")
    now = datetime.now(timezone.utc)
    item["menu_item_id"] = gen_id("menu")
    item["name"] = f"{item.get('name', 'Item')} (Copy)"
    item["available_today"] = False
    item["created_at"] = now
    item["updated_at"] = now
    await db.menu_items.insert_one(item)
    item.pop("_id", None)
    for k in ("created_at", "updated_at"):
        item[k] = item[k].isoformat()
    return item


@api.delete("/ops/menu/{menu_item_id}")
async def ops_delete_menu_item(menu_item_id: str, request: Request):
    await require_permission(request, "manage_menu")
    result = await db.menu_items.delete_one({"menu_item_id": menu_item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Menu item deleted"}


@api.put("/ops/menu/{menu_item_id}/availability")
async def ops_toggle_availability(menu_item_id: str, body: AvailabilityBody, request: Request):
    await require_permission(request, "manage_menu")
    result = await db.menu_items.update_one(
        {"menu_item_id": menu_item_id},
        {"$set": {"available_today": body.available_today, "updated_at": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Menu item not found")
    return {"message": "Updated", "available_today": body.available_today}


@api.get("/ops/orders")
async def ops_list_orders(request: Request, range: Optional[str] = None, vendor_id: Optional[str] = None,
                          status: Optional[str] = None, page: int = 1, page_size: int = 25):
    me = await require_permission(request, "view_orders")
    cfg = await get_settings_doc()
    query: dict = {}
    now = datetime.now(timezone.utc)
    if range == "today":
        query["created_at"] = {"$gte": _day_start(now)}
    elif range == "week":
        query["created_at"] = {"$gte": _day_start(now) - timedelta(days=7)}
    if vendor_id:
        query["vendor_id"] = vendor_id
    if status:
        query["status"] = status
    total = await db.orders.count_documents(query)
    skip = max(page - 1, 0) * page_size
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for o in orders:
        o["commission"] = round(order_revenue(o) * cfg["commission_rate"], 2)
        o["order_value"] = o.get("total_amount", 0)
        o["customer_name"] = o.get("user_name", "Customer")
        # Pickup codes are only exposed to admins in the ops list.
        if me.get("role") != "admin":
            o.pop("pickup_code", None)
        if isinstance(o.get("created_at"), datetime):
            o["created_at"] = o["created_at"].isoformat()
    return {"items": orders, "total": total, "page": page, "page_size": page_size}


@api.put("/ops/orders/{order_id}/status")
async def ops_update_order_status(order_id: str, body: dict, request: Request):
    await require_permission(request, "update_order_status")
    status = body.get("status")
    if status not in ("reserved", "picked_up", "cancelled", "expired", "refunded"):
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await db.orders.update_one({"order_id": order_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Status updated", "status": status}

@api.post("/ops/orders/verify-pickup")
async def ops_verify_pickup(body: dict, request: Request):
    """Ops-side pickup code verification. Staff enter a customer's 6-digit code to
    confirm it on the spot. A valid, still-reserved order is completed (marked
    picked_up); already-handled orders return their current state with a clear message."""
    me = await require_permission(request, "update_order_status")
    code = (body.get("code") or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="Please enter a pickup code.")

    order = await db.orders.find_one({"pickup_code": code}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="No order found for this pickup code.")

    def summary(o: dict, ok: bool, message: str) -> dict:
        created = o.get("created_at")
        return {
            "valid": ok,
            "message": message,
            "order_id": o.get("order_id"),
            "status": o.get("status"),
            "customer_name": o.get("user_name", "Customer"),
            "vendor_name": o.get("vendor_name", "—"),
            "food_item_name": o.get("food_item_name", "—"),
            "quantity": o.get("quantity", 1),
            "order_value": o.get("total_amount", 0),
            "pickup_start_time": o.get("pickup_start_time", ""),
            "pickup_end_time": o.get("pickup_end_time", ""),
            "created_at": created.isoformat() if isinstance(created, datetime) else created,
        }

    status = order.get("status")
    if status == "picked_up":
        return summary(order, False, "This order was already picked up.")
    if status in ("cancelled", "refunded", "expired"):
        return summary(order, False, f"This order is {status} and can no longer be verified.")

    now = datetime.now(timezone.utc)
    result = await db.orders.update_one(
        {"pickup_code": code, "status": "reserved"},
        {"$set": {
            "status": "picked_up",
            "pickup_verified": True,
            "pickup_verified_at": now,
            "pickup_verified_by": me.get("name") or me.get("email") or "Ops",
        }},
    )
    if result.modified_count == 0:
        fresh = await db.orders.find_one({"pickup_code": code}, {"_id": 0}) or order
        return summary(fresh, False, "This order could not be completed. Please refresh and try again.")

    order["status"] = "picked_up"
    return summary(order, True, "Pickup verified — order marked as picked up.")




@api.post("/ops/orders/{order_id}/refund")
async def ops_refund_order(order_id: str, request: Request):
    """Mark an order refunded and invalidate its pickup code so it can no longer
    be verified. Does not trigger a money refund in Razorpay (do that from the
    Razorpay dashboard; the refund webhook will also mark orders refunded)."""
    user = await require_permission(request, "manage_payouts")
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.get("status") == "refunded":
        return {"message": "Order already refunded", "status": "refunded"}
    await db.orders.update_one({"order_id": order_id}, {"$set": {
        "status": "refunded",
        "pickup_code": None,
        "pickup_verified": False,
        "refunded_at": datetime.now(timezone.utc),
        "refunded_by": user.get("name") or user.get("email") or "Staff",
    }})
    # Restore surplus quantity if it was decremented and not yet picked up.
    if order.get("order_type") == "surplus" and order.get("status") == "reserved" and order.get("quantity"):
        await db.menu_items.update_one(
            {"menu_item_id": order.get("food_item_id"), "quantity_available": {"$ne": None}},
            {"$inc": {"quantity_available": order.get("quantity", 0)}},
        )
    return {"message": "Order refunded", "status": "refunded"}


class TestOrderBody(BaseModel):
    vendor_id: Optional[str] = None


@api.post("/ops/orders/test")
async def ops_create_test_order(body: TestOrderBody, request: Request):
    """Admin-only helper: insert a PAID/reserved order with a real pickup code so
    vendor pickup verification can be tested without going through Razorpay."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendor = None
    if body.vendor_id:
        vendor = await db.vendors.find_one({"vendor_id": body.vendor_id}, {"_id": 0})
    if not vendor:
        vendor = await db.vendors.find_one({"status": "active"}, {"_id": 0}) or await db.vendors.find_one({}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="No vendor found. Create a vendor first.")
    item = await db.menu_items.find_one({"vendor_id": vendor["vendor_id"]}, {"_id": 0})
    now = datetime.now(timezone.utc)
    code = await _gen_pickup_code()
    order_id = gen_id("order")
    price = (item or {}).get("original_price") or 100
    order_doc = {
        "order_id": order_id,
        "user_id": user["user_id"],
        "user_name": "Test Customer (Ops)",
        "food_item_id": (item or {}).get("menu_item_id", "test_item"),
        "food_item_name": (item or {}).get("name", "Test Item"),
        "vendor_id": vendor["vendor_id"],
        "vendor_name": vendor.get("name", ""),
        "quantity": 1,
        "order_type": "takeaway",
        "discounted_price": price,
        "item_subtotal": price,
        "total_amount": round(price * 1.05, 2),
        "status": "reserved",
        "pickup_code": code,
        "pickup_verified": False,
        "pickup_verified_at": None,
        "pickup_verified_by": None,
        "payment_confirmed_at": now,
        "pickup_start_time": vendor.get("pickup_start_time", "18:00"),
        "pickup_end_time": vendor.get("pickup_end_time", "21:00"),
        "razorpay_order_id": f"test_{order_id}",
        "razorpay_payment_id": f"test_pay_{order_id}",
        "is_test": True,
        "created_at": now,
    }
    await db.orders.insert_one(order_doc)
    return {"order_id": order_id, "pickup_code": code, "vendor_name": vendor.get("name", ""), "vendor_id": vendor["vendor_id"]}


# ── Admin testing utilities (create / list / delete labelled test orders) ──

def _test_order_view(o: dict) -> dict:
    ct = o.get("created_at")
    return {
        "order_id": o.get("order_id"),
        "user_name": o.get("user_name"),
        "vendor_name": o.get("vendor_name"),
        "food_item_name": o.get("food_item_name"),
        "quantity": o.get("quantity", 1),
        "total_amount": o.get("total_amount", 0),
        "status": o.get("status"),
        "pickup_code": o.get("pickup_code"),
        "created_at": ct.isoformat() if hasattr(ct, "isoformat") else ct,
    }


@api.post("/ops/testing/orders")
async def ops_testing_create_order(request: Request):
    """Admin-only: insert a clearly-labelled PAID test order for the vendor
    'Perfectly Good' (falls back to any active vendor) so flows can be tested."""
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    vendor = await db.vendors.find_one({"name": {"$regex": "^perfectly good$", "$options": "i"}}, {"_id": 0})
    if not vendor:
        vendor = await db.vendors.find_one({"status": "active"}, {"_id": 0}) or await db.vendors.find_one({}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="No vendor found. Create a vendor first.")
    item = await db.menu_items.find_one({"vendor_id": vendor["vendor_id"]}, {"_id": 0})
    now = datetime.now(timezone.utc)
    code = await _gen_pickup_code()
    order_id = gen_id("order")
    price = (item or {}).get("original_price") or 100
    order_doc = {
        "order_id": order_id,
        "user_id": user["user_id"],
        "user_name": "Test Customer",
        "food_item_id": (item or {}).get("menu_item_id", "test_item"),
        "food_item_name": (item or {}).get("name", "Test Item"),
        "vendor_id": vendor["vendor_id"],
        "vendor_name": vendor.get("name", ""),
        "quantity": 1,
        "order_type": "takeaway",
        "discounted_price": price,
        "item_subtotal": price,
        "total_amount": round(price * 1.05, 2),
        "status": "reserved",
        "pickup_code": code,
        "pickup_verified": False,
        "pickup_verified_at": None,
        "pickup_verified_by": None,
        "payment_confirmed_at": now,
        "pickup_start_time": vendor.get("pickup_start_time", "18:00"),
        "pickup_end_time": vendor.get("pickup_end_time", "21:00"),
        "razorpay_order_id": f"test_{order_id}",
        "razorpay_payment_id": f"test_pay_{order_id}",
        "is_test": True,
        "created_at": now,
    }
    await db.orders.insert_one(order_doc)
    return _test_order_view(order_doc)


@api.get("/ops/testing/orders")
async def ops_testing_list_orders(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    rows = await db.orders.find({"is_test": True}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"items": [_test_order_view(o) for o in rows], "total": len(rows)}


@api.delete("/ops/testing/orders/{order_id}")
async def ops_testing_delete_order(order_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    res = await db.orders.delete_one({"order_id": order_id, "is_test": True})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Test order not found")
    return {"message": "Test order deleted"}


@api.get("/ops/users")
async def ops_list_users(request: Request, search: Optional[str] = None, page: int = 1, page_size: int = 25):
    await require_permission(request, "view_users")
    query: dict = {"role": "user"}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
        ]
    total = await db.users.count_documents(query)
    skip = max(page - 1, 0) * page_size
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    # aggregate orders + money saved
    uids = [u["user_id"] for u in users]
    orders = await db.orders.find({"user_id": {"$in": uids}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(100000)
    items = await db.menu_items.find({}, {"_id": 0, "menu_item_id": 1, "original_price": 1}).to_list(100000)
    orig_map = {i["menu_item_id"]: i.get("original_price", 0) for i in items}
    by_user: dict = {}
    for o in orders:
        d = by_user.setdefault(o.get("user_id"), {"orders": 0, "saved": 0.0})
        d["orders"] += 1
        orig = orig_map.get(o.get("food_item_id"), 0) * o.get("quantity", 1)
        d["saved"] += max(orig - order_revenue(o), 0)
    for u in users:
        d = by_user.get(u["user_id"], {"orders": 0, "saved": 0})
        u["orders"] = d["orders"]
        u["money_saved"] = round(d["saved"], 2)
        if isinstance(u.get("created_at"), datetime):
            u["created_at"] = u["created_at"].isoformat()
    return {"items": users, "total": total, "page": page, "page_size": page_size}


async def _compute_payouts(period: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None):
    cfg = await get_settings_doc()
    now = datetime.now(timezone.utc)
    order_q: dict = {"status": "picked_up"}
    if period == "weekly":
        order_q["created_at"] = {"$gte": _day_start(now) - timedelta(days=7)}
    elif period == "monthly":
        order_q["created_at"] = {"$gte": _month_start(now)}
    elif start and end:
        try:
            order_q["created_at"] = {"$gte": datetime.fromisoformat(start), "$lte": datetime.fromisoformat(end)}
        except Exception:
            pass
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(2000)
    completed = await db.orders.find(order_q, {"_id": 0}).to_list(200000)
    all_completed = await db.orders.find({"status": "picked_up"}, {"_id": 0}).to_list(200000)
    pending_q = await db.orders.find({"status": "reserved"}, {"_id": 0, "vendor_id": 1}).to_list(200000)
    rev_by_vendor: dict = {}
    for o in completed:
        rev_by_vendor.setdefault(o.get("vendor_id"), {"rev": 0.0, "count": 0})
        rev_by_vendor[o.get("vendor_id")]["rev"] += order_revenue(o)
        rev_by_vendor[o.get("vendor_id")]["count"] += 1
    all_net_by_vendor: dict = {}
    for o in all_completed:
        all_net_by_vendor.setdefault(o.get("vendor_id"), 0.0)
        all_net_by_vendor[o.get("vendor_id")] += order_revenue(o)
    pending_orders_by_vendor: dict = {}
    for o in pending_q:
        pending_orders_by_vendor[o.get("vendor_id")] = pending_orders_by_vendor.get(o.get("vendor_id"), 0) + 1
    payouts = await db.payouts.find({}, {"_id": 0}).to_list(200000)
    paid_by_vendor: dict = {}
    last_payout: dict = {}
    for p in payouts:
        vid = p.get("vendor_id")
        paid_by_vendor[vid] = paid_by_vendor.get(vid, 0.0) + p.get("amount", 0)
        ca = p.get("created_at")
        if isinstance(ca, datetime):
            if vid not in last_payout or ca > last_payout[vid]:
                last_payout[vid] = ca
    result = []
    for v in vendors:
        vid = v["vendor_id"]
        rv = rev_by_vendor.get(vid, {"rev": 0.0, "count": 0})
        total_sales = round(rv["rev"], 2)
        commission = round(total_sales * cfg["commission_rate"], 2)
        gst_on_comm = round(commission * cfg["gst_on_commission"], 2)
        net_payable = round(total_sales - commission - gst_on_comm, 2)
        all_net = round(all_net_by_vendor.get(vid, 0.0) * (1 - cfg["commission_rate"] * (1 + cfg["gst_on_commission"])), 2)
        paid = round(paid_by_vendor.get(vid, 0.0), 2)
        pending_amt = round(max(all_net - paid, 0), 2)
        result.append({
            "vendor_id": vid, "vendor_name": v.get("name", ""),
            "total_sales": total_sales, "commission": commission, "gst_on_commission": gst_on_comm,
            "net_payable": net_payable, "completed_orders": rv["count"],
            "pending_orders": pending_orders_by_vendor.get(vid, 0),
            "last_payout_date": last_payout[vid].isoformat() if vid in last_payout else None,
            "total_paid": paid, "pending_payout": pending_amt,
            "status": "paid" if pending_amt <= 0 and paid > 0 else ("pending" if pending_amt > 0 else "no_dues"),
        })
    result.sort(key=lambda r: r["pending_payout"], reverse=True)
    return result


@api.get("/ops/payouts")
async def ops_payouts(request: Request, period: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None):
    await require_permission(request, "view_finance")
    return await _compute_payouts(period, start, end)


@api.get("/ops/payouts/{vendor_id}/history")
async def ops_payout_history(vendor_id: str, request: Request):
    await require_permission(request, "view_finance")
    payouts = await db.payouts.find({"vendor_id": vendor_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for p in payouts:
        if isinstance(p.get("created_at"), datetime):
            p["created_at"] = p["created_at"].isoformat()
    return payouts


@api.post("/ops/payouts/mark-paid")
async def ops_mark_paid(body: MarkPaidBody, request: Request):
    user = await require_permission(request, "manage_payouts")
    vendor = await db.vendors.find_one({"vendor_id": body.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    payout_doc = {
        "payout_id": gen_id("payout"), "vendor_id": body.vendor_id,
        "vendor_name": vendor.get("name", ""), "amount": round(body.amount, 2),
        "reference_number": body.reference_number or "", "notes": body.notes or "",
        "method": body.method or "bank_transfer", "status": "paid",
        "paid_by": user.get("name", "Staff"), "created_at": datetime.now(timezone.utc),
    }
    await db.payouts.insert_one(payout_doc)
    payout_doc.pop("_id", None)
    payout_doc["created_at"] = payout_doc["created_at"].isoformat()
    return payout_doc


@api.get("/ops/payment-failures")
async def ops_payment_failures(request: Request, page: int = 1, page_size: int = 50):
    """Failed Razorpay payments logged by the webhook (finance visibility)."""
    await require_permission(request, "view_finance")
    total = await db.payment_failures.count_documents({})
    skip = max(page - 1, 0) * page_size
    rows = await db.payment_failures.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for r in rows:
        if isinstance(r.get("created_at"), datetime):
            r["created_at"] = r["created_at"].isoformat()
    return {"items": rows, "total": total, "page": page, "page_size": page_size}


@api.get("/ops/settings")
async def ops_get_settings(request: Request):
    await require_permission(request, "view_dashboard")
    cfg = await get_settings_doc()
    cfg.pop("_id", None)
    return cfg


@api.put("/ops/settings")
async def ops_update_settings(body: dict, request: Request):
    await require_permission(request, "manage_settings")
    allowed = set(DEFAULT_SETTINGS.keys())
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.settings.update_one({"_id": "platform"}, {"$set": updates}, upsert=True)
    cfg = await get_settings_doc()
    cfg.pop("_id", None)
    return cfg


@api.get("/ops/roles")
async def ops_roles(request: Request):
    await require_permission(request, "view_dashboard")
    return {"permissions": PERMISSIONS, "roles": {r: sorted(p) for r, p in ROLE_PERMISSIONS.items() if r in STAFF_ROLES}}


@api.get("/ops/assignable-ops")
async def ops_assignable(request: Request):
    """Operations staff that a vendor can be assigned to (admin uses this for the picker)."""
    await require_permission(request, "manage_vendors")
    staff = await db.users.find({"role": "operations"}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).to_list(500)
    return [{"user_id": s["user_id"], "name": s.get("name") or s.get("email") or "", "email": s.get("email", "")} for s in staff]


@api.get("/ops/staff")
async def ops_list_staff(request: Request):
    await require_permission(request, "manage_roles")
    staff = await db.users.find({"role": {"$in": list(STAFF_ROLES)}}, {"_id": 0, "password_hash": 0}).to_list(500)
    for s in staff:
        if isinstance(s.get("created_at"), datetime):
            s["created_at"] = s["created_at"].isoformat()
        s["permissions"] = sorted(get_effective_permissions(s.get("role", "user"), s.get("permission_overrides")))
    return staff


@api.post("/ops/staff")
async def ops_create_staff(body: StaffBody, request: Request):
    await require_permission(request, "manage_roles")
    if body.role not in STAFF_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    if len((body.password or "").strip()) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    email = body.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "user_id": gen_id("user"), "email": email, "name": body.name,
        "password_hash": hash_password(body.password.strip()),
        "role": body.role, "permission_overrides": body.permission_overrides or {},
        "picture": None, "location": None, "created_at": datetime.now(timezone.utc),
    }
    await db.users.insert_one(doc)
    return {"message": "Staff created", "user_id": doc["user_id"], "email": email}


@api.put("/ops/staff/{user_id}/role")
async def ops_update_staff_role(user_id: str, body: dict, request: Request):
    await require_permission(request, "manage_roles")
    role = body.get("role")
    if role not in STAFF_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    updates = {"role": role}
    if "permission_overrides" in body:
        updates["permission_overrides"] = body["permission_overrides"] or {}
    result = await db.users.update_one({"user_id": user_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Role updated", "role": role}


@api.delete("/ops/staff/{user_id}")
async def ops_delete_staff(user_id: str, request: Request):
    me = await require_permission(request, "manage_roles")
    if me["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target or target.get("role") not in STAFF_ROLES:
        raise HTTPException(status_code=404, detail="Staff member not found")
    await db.users.delete_one({"user_id": user_id})
    return {"message": "Staff removed"}


@api.put("/ops/vendors/{vendor_id}/password")
async def ops_set_vendor_password(vendor_id: str, body: dict, request: Request):
    """Admin-only: set a new (temporary) password on a vendor's login account
    directly, without needing the old password."""
    user = await require_permission(request, "manage_vendors")
    _admin_only(user)
    new_password = (body.get("password") or "").strip()
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    uid = vendor.get("user_id")
    if not uid:
        raise HTTPException(status_code=400, detail="This vendor has no linked login account")
    res = await db.users.update_one({"user_id": uid}, {"$set": {"password_hash": hash_password(new_password)}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor login account not found")
    return {"message": "Temporary password set"}


@api.put("/ops/vendors/{vendor_id}/email")
async def ops_set_vendor_email(vendor_id: str, body: dict, request: Request):
    """Admin-only: change a vendor's email (updates both the vendor record and its login account)."""
    user = await require_permission(request, "manage_vendors")
    _admin_only(user)
    email = (body.get("email") or "").strip().lower()
    if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
        raise HTTPException(status_code=400, detail="Please enter a valid email address")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    uid = vendor.get("user_id")
    clash = await db.users.find_one({"email": email, "user_id": {"$ne": uid}}, {"_id": 0})
    if clash:
        raise HTTPException(status_code=400, detail="That email is already in use by another account")
    now = datetime.now(timezone.utc)
    await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": {"email": email, "updated_at": now}})
    if uid:
        await db.users.update_one({"user_id": uid}, {"$set": {"email": email}})
    return {"message": "Vendor email updated", "email": email}


@api.post("/ops/vendors/{vendor_id}/menu/bulk-delete")
async def ops_bulk_delete_menu(vendor_id: str, body: dict, request: Request):
    """Delete multiple menu items for a vendor in one action."""
    user = await require_permission(request, "manage_vendors")
    v = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if user.get("role") == "operations" and v.get("assigned_ops") != user["user_id"]:
        raise HTTPException(status_code=403, detail="This vendor is not assigned to you")
    ids = body.get("menu_item_ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="No menu items selected")
    res = await db.menu_items.delete_many({"vendor_id": vendor_id, "menu_item_id": {"$in": ids}})
    return {"message": f"Deleted {res.deleted_count} item(s)", "deleted": res.deleted_count}


@api.put("/ops/staff/{user_id}/password")
async def ops_set_staff_password(user_id: str, body: dict, request: Request):
    await require_permission(request, "manage_roles")
    new_password = (body.get("password") or "").strip()
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target or target.get("role") not in STAFF_ROLES:
        raise HTTPException(status_code=404, detail="Staff member not found")
    await db.users.update_one({"user_id": user_id}, {"$set": {"password_hash": hash_password(new_password)}})
    return {"message": "Password updated"}


@api.get("/ops/search")
async def ops_search(request: Request, q: str):
    await require_permission(request, "view_vendors")
    rx = {"$regex": q, "$options": "i"}
    vendors = await db.vendors.find({"$or": [{"name": rx}, {"email": rx}, {"phone": rx}]}, {"_id": 0, "vendor_id": 1, "name": 1, "category": 1, "status": 1}).limit(10).to_list(10)
    customers = await db.users.find({"role": "user", "$or": [{"name": rx}, {"email": rx}, {"phone": rx}]}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).limit(10).to_list(10)
    items = await db.menu_items.find({"name": rx}, {"_id": 0, "menu_item_id": 1, "name": 1, "vendor_id": 1}).limit(10).to_list(10)
    orders = await db.orders.find({"$or": [{"order_id": rx}, {"user_name": rx}, {"vendor_name": rx}]}, {"_id": 0, "order_id": 1, "user_name": 1, "vendor_name": 1, "status": 1}).limit(10).to_list(10)
    return {"vendors": vendors, "customers": customers, "menu_items": items, "orders": orders}


@api.post("/ops/upload")
async def ops_upload(file: UploadFile = File(...), request: Request = None):
    if request:
        await require_permission(request, "upload_images")
    contents = await file.read()
    import base64 as _b64
    mime = file.content_type or "image/jpeg"
    data_uri = f"data:{mime};base64,{_b64.b64encode(contents).decode('utf-8')}"
    return {"url": data_uri}


# ══════════════════════════════════════════════════════════════════════════
#  PHASE 2 — AI Import, Bulk Import, Exports, Analytics, Vendor Performance
# ══════════════════════════════════════════════════════════════════════════

def _as_dt(v) -> datetime:
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    try:
        d = datetime.fromisoformat(str(v))
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except Exception:
        return datetime.now(timezone.utc)


def _coerce_price(v):
    try:
        return float(str(v).replace("₹", "").replace(",", "").strip())
    except Exception:
        return None


def _parse_json_items(text: str) -> list:
    import json, re
    if not text:
        return []
    t = str(text).strip()
    t = re.sub(r"^```(json)?", "", t).strip()
    t = re.sub(r"```$", "", t).strip()
    s, e = t.find("["), t.rfind("]")
    if s != -1 and e != -1:
        t = t[s:e + 1]
    try:
        data = json.loads(t)
    except Exception:
        return []
    out = []
    for d in (data if isinstance(data, list) else []):
        if not isinstance(d, dict):
            continue
        name = str(d.get("name") or "").strip()
        if not name:
            continue
        out.append({
            "name": name, "description": str(d.get("description") or ""),
            "original_price": _coerce_price(d.get("original_price")) or 0,
            "discounted_price": None, "food_type": "veg", "contains_egg": False,
            "serving_size": "", "category": "", "available_today": False,
        })
    return out


def _normalize_menu_row(d: dict) -> dict:
    dd = {(str(k).strip().lower() if k else ""): v for k, v in d.items()}
    g = lambda *keys: next((dd[k] for k in keys if k in dd and dd[k] not in (None, "")), None)

    def find(*subs):
        # Match the first non-empty column whose header contains any of the substrings
        for k, v in dd.items():
            if v in (None, ""):
                continue
            if any(s in k for s in subs):
                return v
        return None

    truthy = lambda v: str(v).strip().lower() in ("yes", "true", "1", "y", "available", "live")

    veg_val = find("veg", "food type", "food_type")  # 'veg', 'veg/non-veg', 'non veg', etc.
    if veg_val is None:
        veg_is = True
    else:
        vv = str(veg_val).strip().lower()
        veg_is = not ("non" in vv or vv in ("no", "n", "false", "0", "nonveg", "non-veg", "non_veg"))

    return {
        "name": str(g("item name", "name", "item") or find("item", "dish", "name") or "").strip(),
        "description": str(g("description", "desc") or ""),
        "original_price": _coerce_price(g("original price", "price", "mrp") or find("price")) or 0,
        "discounted_price": _coerce_price(g("discounted price", "discount price", "sale price")),
        "serving_size": str(g("serving size", "serving") or ""),
        "category": str(g("category") or ""),
        "food_type": "veg" if veg_is else "non_veg",
        "contains_egg": truthy(find("egg") or ""),
        "available_today": truthy(find("available today", "available", "live") or ""),
    }


@api.post("/ops/menu-import/parse-file")
async def menu_import_parse_file(file: UploadFile = File(...), request: Request = None):
    if request:
        await require_permission(request, "manage_menu")
    content = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    if fname.endswith(".csv"):
        import csv, io
        text = content.decode("utf-8-sig", errors="ignore")
        for r in csv.DictReader(io.StringIO(text)):
            rows.append(_normalize_menu_row(r))
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1), [])
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            rows.append(_normalize_menu_row(d))
    else:
        raise HTTPException(status_code=400, detail="Unsupported file. Upload a CSV or XLSX file.")
    rows = [r for r in rows if r.get("name")]
    return {"items": rows, "count": len(rows)}


@api.post("/ops/vendors/{vendor_id}/menu/bulk")
async def ops_bulk_add_menu(vendor_id: str, body: dict, request: Request):
    await require_permission(request, "manage_menu")
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    cfg = await get_settings_doc()
    now = datetime.now(timezone.utc)
    docs = []
    for it in body.get("items", []):
        op = _coerce_price(it.get("original_price")) or 0
        name = str(it.get("name") or "").strip()
        if not name or op <= 0:
            continue
        dp = it.get("discounted_price")
        dp = _coerce_price(dp) if dp not in (None, "") else None
        docs.append({
            "menu_item_id": gen_id("menu"), "vendor_id": vendor_id, "name": name,
            "description": str(it.get("description") or ""), "original_price": op,
            "discounted_price": dp, "category": str(it.get("category") or vendor.get("category", "")),
            "serving_size": str(it.get("serving_size") or ""), "food_type": it.get("food_type") or "veg",
            "contains_egg": bool(it.get("contains_egg")), "available_today": False,
            "in_stock": True, "quantity_available": None,
            "image_url": it.get("image_url") or "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=600",
            "created_at": now, "updated_at": now,
        })
    if docs:
        await db.menu_items.insert_many(docs)
    return {"created": len(docs)}


async def _export_dataset(entity: str, user: Optional[dict] = None):
    cfg = await get_settings_doc()
    if entity == "vendors":
        query: dict = {}
        # Operations staff export only the vendors assigned to them.
        if user and user.get("role") == "operations":
            query["assigned_ops"] = user["user_id"]
        vendors = await db.vendors.find(query, {"_id": 0}).to_list(5000)
        agg = await _vendor_aggregates([v["vendor_id"] for v in vendors])
        ops_names = await _ops_name_map()
        header = ["Vendor", "Owner", "Category", "Phone", "Email", "Service Type", "Status",
                  "Discount %", "Assigned Ops", "Menu Items", "Orders", "Revenue", "Date Added"]
        rows = [[v.get("name"), v.get("owner_name"), v.get("category"), v.get("phone"), v.get("email"),
                 v.get("service_type"), v.get("status"), v.get("discount_percentage", 0) or 0,
                 ops_names.get(v.get("assigned_ops", ""), "") or "Unassigned",
                 agg.get(v["vendor_id"], {}).get("menu_count", 0),
                 agg.get(v["vendor_id"], {}).get("order_count", 0), agg.get(v["vendor_id"], {}).get("revenue", 0),
                 fmt_dt(v.get("created_at"))] for v in vendors]
        return header, rows
    if entity == "orders":
        orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(100000)
        header = ["Order ID", "Customer", "Vendor", "Item", "Qty", "Order Value", "Commission", "Status", "Payment", "Created"]
        rows = [[o.get("order_id"), o.get("user_name"), o.get("vendor_name"), o.get("food_item_name"),
                 o.get("quantity"), o.get("total_amount"), round(order_revenue(o) * cfg["commission_rate"], 2),
                 o.get("status"), "Paid" if o.get("razorpay_payment_id") else "Pending", fmt_dt(o.get("created_at"))] for o in orders]
        return header, rows
    if entity == "customers":
        users = await db.users.find({"role": "user"}, {"_id": 0, "password_hash": 0}).to_list(100000)
        uids = [u["user_id"] for u in users]
        orders = await db.orders.find({"user_id": {"$in": uids}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(200000)
        items = await db.menu_items.find({}, {"_id": 0, "menu_item_id": 1, "original_price": 1}).to_list(200000)
        orig = {i["menu_item_id"]: i.get("original_price", 0) for i in items}
        from collections import defaultdict
        agg = defaultdict(lambda: {"orders": 0, "saved": 0.0})
        for o in orders:
            a = agg[o.get("user_id")]; a["orders"] += 1
            a["saved"] += max(orig.get(o.get("food_item_id"), 0) * o.get("quantity", 1) - order_revenue(o), 0)
        header = ["Name", "Email", "Phone", "Orders", "Money Saved", "Joined"]
        rows = [[u.get("name"), u.get("email"), u.get("phone"), agg[u["user_id"]]["orders"],
                 round(agg[u["user_id"]]["saved"], 2), fmt_dt(u.get("created_at"))] for u in users]
        return header, rows
    if entity == "menu":
        items = await db.menu_items.find({}, {"_id": 0}).to_list(200000)
        vmap = {v["vendor_id"]: v.get("name", "") for v in await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "name": 1}).to_list(5000)}
        header = ["Vendor", "Item", "Description", "Category", "Original Price", "Discounted Price", "Veg/Non-Veg", "Contains Egg", "Serving Size", "Available Today"]
        rows = [[vmap.get(m.get("vendor_id"), ""), m.get("name"), m.get("description"), m.get("category"),
                 m.get("original_price"), m.get("discounted_price"), m.get("food_type"),
                 "Yes" if m.get("contains_egg") else "No", m.get("serving_size"), "Yes" if m.get("available_today") else "No"] for m in items]
        return header, rows
    if entity == "payouts":
        data = await _compute_payouts()
        header = ["Vendor", "Total Sales", "Commission", "GST on Commission", "Net Payable", "Completed Orders", "Pending Orders", "Total Paid", "Pending Payout", "Status", "Last Payout"]
        rows = [[d["vendor_name"], d["total_sales"], d["commission"], d["gst_on_commission"], d["net_payable"],
                 d["completed_orders"], d["pending_orders"], d["total_paid"], d["pending_payout"], d["status"],
                 fmt_dt(d.get("last_payout_date"))] for d in data]
        return header, rows
    raise HTTPException(status_code=404, detail="Unknown export entity")


def fmt_dt(v):
    if not v:
        return ""
    try:
        return _as_dt(v).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(v)


@api.get("/ops/export/{entity}")
async def ops_export(entity: str, request: Request, format: str = "csv"):
    perm_map = {"vendors": "view_vendors", "orders": "view_orders", "customers": "view_users", "menu": "view_vendors", "payouts": "view_finance"}
    if entity not in perm_map:
        raise HTTPException(status_code=404, detail="Unknown export entity")
    await require_permission(request, perm_map[entity])
    from fastapi.responses import StreamingResponse
    user = await get_current_user(request)
    header, rows = await _export_dataset(entity, user)
    if format == "xlsx":
        import openpyxl, io
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = entity[:31]
        ws.append(header)
        for r in rows:
            ws.append(["" if c is None else c for c in r])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f'attachment; filename="{entity}.xlsx"'})
    import csv, io
    buf = io.StringIO(); w = csv.writer(buf); w.writerow(header); w.writerows(rows)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{entity}.csv"'})


@api.get("/ops/analytics")
async def ops_analytics(request: Request, days: int = 30):
    await require_permission(request, "view_dashboard")
    from collections import defaultdict
    cfg = await get_settings_doc()
    now = datetime.now(timezone.utc)
    days = max(7, min(days, 90))
    start = _day_start(now) - timedelta(days=days - 1)

    window_orders = await db.orders.find({"created_at": {"$gte": start}, "status": {"$nin": ["cancelled", "refunded"]}}, {"_id": 0}).to_list(200000)
    rev_day, ord_day = defaultdict(float), defaultdict(int)
    for o in window_orders:
        key = _as_dt(o.get("created_at")).strftime("%Y-%m-%d")
        rev_day[key] += order_revenue(o); ord_day[key] += 1
    trend = []
    for i in range(days):
        d = (start + timedelta(days=i)).strftime("%Y-%m-%d")
        trend.append({"date": d, "revenue": round(rev_day.get(d, 0), 2), "orders": ord_day.get(d, 0)})

    completed = await db.orders.find({"status": "picked_up"}, {"_id": 0}).to_list(200000)
    items = await db.menu_items.find({}, {"_id": 0, "menu_item_id": 1, "original_price": 1, "category": 1, "vendor_id": 1}).to_list(200000)
    orig = {i["menu_item_id"]: i.get("original_price", 0) for i in items}
    cat_by_item = {i["menu_item_id"]: (i.get("category") or "") for i in items}
    item_vendor = {i["menu_item_id"]: i.get("vendor_id") for i in items}
    vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "name": 1, "category": 1}).to_list(5000)
    vcat = {v["vendor_id"]: v.get("category", "") for v in vendors}

    total_rev = round(sum(order_revenue(o) for o in completed), 2)
    money_saved = 0.0; food_value = 0.0
    for o in completed:
        ov = orig.get(o.get("food_item_id"), 0) * o.get("quantity", 1)
        food_value += ov; money_saved += max(ov - order_revenue(o), 0)

    rev_vendor = defaultdict(lambda: {"rev": 0.0, "orders": 0, "name": ""})
    item_sales = defaultdict(lambda: {"qty": 0, "rev": 0.0, "name": ""})
    cat_agg = defaultdict(lambda: {"orders": 0, "revenue": 0.0})
    for o in completed:
        rv = rev_vendor[o.get("vendor_id")]; rv["rev"] += order_revenue(o); rv["orders"] += 1; rv["name"] = o.get("vendor_name", "")
        si = item_sales[o.get("food_item_id")]; si["qty"] += o.get("quantity", 1); si["rev"] += order_revenue(o); si["name"] = o.get("food_item_name", "")
        c = cat_by_item.get(o.get("food_item_id")) or vcat.get(item_vendor.get(o.get("food_item_id"), ""), "") or "Uncategorized"
        cat_agg[c]["orders"] += 1; cat_agg[c]["revenue"] += order_revenue(o)

    top_vendors = sorted([{"vendor_id": k, "name": v["name"], "revenue": round(v["rev"], 2), "orders": v["orders"]} for k, v in rev_vendor.items()], key=lambda x: x["revenue"], reverse=True)[:5]
    top_items = sorted([{"item_id": k, "name": v["name"], "qty": v["qty"], "revenue": round(v["rev"], 2)} for k, v in item_sales.items()], key=lambda x: x["qty"], reverse=True)[:5]
    perf = {v["vendor_id"]: {"vendor_id": v["vendor_id"], "name": v.get("name", ""), "revenue": 0.0, "orders": 0} for v in vendors}
    for k, v in rev_vendor.items():
        if k in perf:
            perf[k]["revenue"] = round(v["rev"], 2); perf[k]["orders"] = v["orders"]
    lowest = sorted(perf.values(), key=lambda x: (x["revenue"], x["orders"]))[:5]
    categories = [{"category": k, "orders": v["orders"], "revenue": round(v["revenue"], 2), "commission": round(v["revenue"] * cfg["commission_rate"], 2)} for k, v in sorted(cat_agg.items(), key=lambda x: x[1]["revenue"], reverse=True)]

    return {
        "trend": trend,
        "totals": {
            "revenue": total_rev, "orders": len(completed),
            "commission": round(total_rev * cfg["commission_rate"], 2),
            "money_saved": round(money_saved, 2), "food_value_rescued": round(food_value, 2),
            "aov": round(total_rev / max(len(completed), 1), 2),
            "new_users": await db.users.count_documents({"role": "user", "created_at": {"$gte": start}}),
            "new_vendors": await db.vendors.count_documents({"created_at": {"$gte": start}}),
            "active_vendors": await db.vendors.count_documents({"status": {"$ne": "inactive"}}),
        },
        "top_vendors": top_vendors, "top_items": top_items, "lowest_vendors": lowest, "categories": categories,
    }


@api.get("/ops/vendors/{vendor_id}/performance")
async def ops_vendor_performance(vendor_id: str, request: Request):
    await require_permission(request, "view_vendors")
    from collections import defaultdict
    cfg = await get_settings_doc()
    now = datetime.now(timezone.utc)
    week0 = _day_start(now) - timedelta(days=7)
    month0 = _month_start(now)
    orders = await db.orders.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(100000)
    valid = [o for o in orders if o.get("status") not in ("cancelled", "refunded")]
    completed = [o for o in orders if o.get("status") == "picked_up"]
    revenue = round(sum(order_revenue(o) for o in completed), 2)
    menu = await db.menu_items.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(2000)
    item_qty = defaultdict(int); names = {}
    for o in completed:
        item_qty[o.get("food_item_id")] += o.get("quantity", 1); names[o.get("food_item_id")] = o.get("food_item_name", "")
    best = max(item_qty.items(), key=lambda x: x[1]) if item_qty else None
    last_order = max([_as_dt(o.get("created_at")) for o in orders], default=None)
    return {
        "total_orders": len(valid),
        "orders_week": len([o for o in valid if _as_dt(o.get("created_at")) >= week0]),
        "orders_month": len([o for o in valid if _as_dt(o.get("created_at")) >= month0]),
        "revenue": revenue,
        "commission": round(revenue * cfg["commission_rate"], 2),
        "aov": round(revenue / max(len(completed), 1), 2),
        "total_listings": len(menu),
        "active_listings_today": len([m for m in menu if m.get("available_today")]),
        "best_selling_item": names.get(best[0]) if best else None,
        "best_selling_qty": best[1] if best else 0,
        "last_order_date": last_order.isoformat() if last_order else None,
    }


# ── Health ──────────────────────────────────────────────────────────────
@api.get("/")
async def root():
    return {"message": "Perfectly Good API", "status": "running"}

# ── Seed Data ───────────────────────────────────────────────────────────
async def seed_data():
    # ── Seed core staff accounts (idempotent). Runs on every startup so a fresh
    #    production Atlas DB gets the team accounts; existing accounts are left
    #    untouched (admin-changed passwords survive). ──
    staff_seed = [
        ("Anubhav", "anubhavg@perfectlygood.in", "Anubhavv", "admin"),
        ("Subhash Ramachandra", "subhashramachandraofficial@gmail.com", "123456789", "operations"),
    ]
    for name, email, pwd, role in staff_seed:
        email = email.strip().lower()
        if not await db.users.find_one({"email": email}):
            await db.users.insert_one({
                "user_id": gen_id("user"), "email": email, "name": name,
                "password_hash": hash_password(pwd), "role": role,
                "permission_overrides": {}, "phone": "", "picture": None,
                "location": None, "created_at": datetime.now(timezone.utc),
            })
            logger.info(f"Staff seeded: {email} ({role})")

    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.vendors.create_index("vendor_id", unique=True)
    await db.vendors.create_index("user_id")
    await db.drops.create_index("item_id", unique=True)
    await db.drops.create_index("vendor_id")
    await db.menu_items.create_index("menu_item_id", unique=True)
    await db.menu_items.create_index("vendor_id")
    await db.orders.create_index("order_id", unique=True)
    await db.orders.create_index("user_id")
    await db.orders.create_index("vendor_id")
    logger.info("Database indexes created")

# ── Migration & staff seeding ───────────────────────────────────────────
STAFF_SEED = [
    {"email": "operations@perfectlygood.in", "password": "ops12345", "role": "operations", "name": "Ops Team"},
    {"email": "success@perfectlygood.in", "password": "success12345", "role": "customer_success", "name": "Customer Success"},
    {"email": "finance@perfectlygood.in", "password": "finance12345", "role": "finance", "name": "Finance Team"},
]


async def migrate_v2():
    """Idempotent: unify drops into menu_items, backfill new vendor/menu fields, seed staff."""
    await get_settings_doc()
    now = datetime.now(timezone.utc)

    # 1. Backfill vendor fields
    async for v in db.vendors.find({}):
        updates = {}
        defaults = {
            "status": "active", "service_type": v.get("service_type", "both"),
            "owner_name": "", "restaurant_phone": "", "full_address": v.get("location", {}).get("address", "") if isinstance(v.get("location"), dict) else "",
            "maps_link": v.get("location", {}).get("maps_url", "") if isinstance(v.get("location"), dict) else "",
            "assigned_ops": "", "notes": [], "pickup_start_time": "18:00", "pickup_end_time": "21:00",
            "discount_percentage": 0, "storefront_image": "",
            "created_at": now, "updated_at": now, "last_order_date": None,
        }
        for k, val in defaults.items():
            if k not in v:
                updates[k] = val
        if updates:
            await db.vendors.update_one({"vendor_id": v["vendor_id"]}, {"$set": updates})

    # 2. Backfill menu_item fields + pull discount/qty/pickup from drops
    drops = await db.drops.find({}, {"_id": 0}).to_list(100000)
    drop_by_menu = {}
    drop_itemid_to_menu = {}
    for d in drops:
        mid = d.get("menu_item_id")
        if mid:
            drop_by_menu[mid] = d
            if d.get("item_id"):
                drop_itemid_to_menu[d["item_id"]] = mid
        # set vendor pickup time from drop if vendor still default
        if d.get("pickup_start_time"):
            await db.vendors.update_one(
                {"vendor_id": d.get("vendor_id"), "pickup_start_time": {"$in": [None, "", "18:00"]}},
                {"$set": {"pickup_start_time": d["pickup_start_time"], "pickup_end_time": d.get("pickup_end_time", "21:00")}},
            )

    async for m in db.menu_items.find({}):
        mid = m["menu_item_id"]
        d = drop_by_menu.get(mid)
        updates = {}
        if "discounted_price" not in m:
            if d:
                updates["discounted_price"] = d.get("discounted_price")
            else:
                updates["discounted_price"] = round(m.get("original_price", 0) * 0.6, 2)
        if "available_today" not in m:
            updates["available_today"] = bool(d.get("is_active")) if d else False
        if "quantity_available" not in m:
            updates["quantity_available"] = d.get("quantity_available") if d else None
        if "expiry" not in m:
            updates["expiry"] = d.get("expiry", "") if d else ""
        for k, val in (("food_type", "veg"), ("contains_egg", False), ("serving_size", ""), ("category", "")):
            if k not in m:
                updates[k] = val
        if "created_at" not in m:
            updates["created_at"] = now
        if updates:
            await db.menu_items.update_one({"menu_item_id": mid}, {"$set": updates})

    # 3. Remap legacy orders: food_item_id was a drop item_id -> menu_item_id; store item_subtotal
    if drop_itemid_to_menu:
        async for o in db.orders.find({"food_item_id": {"$in": list(drop_itemid_to_menu.keys())}}):
            new_mid = drop_itemid_to_menu.get(o.get("food_item_id"))
            d = drops and next((x for x in drops if x.get("item_id") == o.get("food_item_id")), None)
            set_fields = {"food_item_id": new_mid, "legacy_item_id": o.get("food_item_id")}
            if d and o.get("item_subtotal") is None:
                set_fields["item_subtotal"] = round(d.get("discounted_price", 0) * o.get("quantity", 1), 2)
                set_fields["discounted_price"] = d.get("discounted_price", 0)
            await db.orders.update_one({"order_id": o["order_id"]}, {"$set": set_fields})

    # 4. Seed staff (idempotent)
    for s in STAFF_SEED:
        existing = await db.users.find_one({"email": s["email"]})
        if not existing:
            await db.users.insert_one({
                "user_id": gen_id("user"), "email": s["email"], "name": s["name"],
                "password_hash": hash_password(s["password"]), "role": s["role"],
                "permission_overrides": {}, "picture": None, "location": None,
                "created_at": now,
            })
            logger.info(f"Staff seeded: {s['email']}")
        elif existing.get("role") not in STAFF_ROLES:
            await db.users.update_one({"email": s["email"]}, {"$set": {"role": s["role"]}})

    # Promote known founder admin if present
    await db.users.update_one({"email": "anubhavg@perfectlygood.in"}, {"$set": {"role": "admin"}})
    logger.info("migrate_v2 complete")


# ── Deal alerts opt-in (warm empty state on onboarding) ─────────────────
class DealAlertBody(BaseModel):
    area: Optional[str] = None


@api.post("/deal-alerts")
async def create_deal_alert(body: DealAlertBody, request: Request):
    """Customer opts in to be notified (via email) when surplus deals go live."""
    user = await get_current_user(request)
    await db.deal_alerts.update_one(
        {"user_id": user["user_id"]},
        {
            "$set": {
                "user_id": user["user_id"],
                "email": user.get("email"),
                "name": user.get("name"),
                "area": body.area,
                "updated_at": datetime.now(timezone.utc),
            },
            "$setOnInsert": {"created_at": datetime.now(timezone.utc)},
        },
        upsert=True,
    )
    return {"message": "You're on the list"}


# ── Sold-out daily reset (midnight IST) ─────────────────────────────────
async def reset_sold_out_items(catch_up: bool = False):
    """Mark all sold-out items available again.
    - Scheduled run (catch_up=False): reset every sold-out item.
    - Startup catch-up (catch_up=True): reset only items marked sold out on a
      previous day, so items a vendor marked sold out *today* stay sold out
      across a server restart."""
    if catch_up:
        q = {"in_stock": False, "$or": [
            {"sold_out_at": {"$exists": False}},
            {"sold_out_at": None},
            {"sold_out_at": {"$lt": today_ist_str()}},
        ]}
    else:
        q = {"in_stock": False}
    res = await db.menu_items.update_many(
        q, {"$set": {"in_stock": True, "sold_out_at": None, "updated_at": datetime.now(timezone.utc)}}
    )
    if res.modified_count:
        logger.info(f"[sold-out reset] {'catch-up ' if catch_up else ''}reset {res.modified_count} item(s) to available")
    return res.modified_count


scheduler = AsyncIOScheduler(timezone=IST)


# ── Startup / Shutdown ──────────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    await seed_data()
    await migrate_v2()
    # Bring back items whose sold-out day has passed while the server was down.
    await reset_sold_out_items(catch_up=True)
    # Reset every sold-out item back to available at 00:00 IST daily.
    scheduler.add_job(reset_sold_out_items, "cron", hour=0, minute=0,
                      id="sold_out_reset", replace_existing=True)
    if not scheduler.running:
        scheduler.start()
    logger.info("Perfectly Good API started")

@app.on_event("shutdown")
async def shutdown():
    if scheduler.running:
        scheduler.shutdown(wait=False)
    mongo_client.close()

# ── Wire up ─────────────────────────────────────────────────────────────
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
